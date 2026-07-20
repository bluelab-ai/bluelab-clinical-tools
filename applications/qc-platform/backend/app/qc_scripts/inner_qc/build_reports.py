#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Phase 6 一站式：合成 markdown 总报告 + 生成 HTML 可视化报告。

原 `merge_qc.py` + `build_report_viewer.py` 合并而来，一条命令跑完两件事：

1. 读所有 `qc_*.json`（含 Phase 3 的 `qc_ext_*.json`）→ 跑跨表规则 R-021/R-029
   → 输出 markdown 总报告
2. 读所有 `qc_*.md` + 原始 `xlsx` → 输出单页自包含 HTML 可视化报告

用法::

    python3 build_reports.py <qc_dir> \\
        --tables-dir ./tables_output \\
        [--baseline ./tables_output/baseline.json] \\
        [--source "<原始docx/pdf>"] \\
        [--md-out  ./qc_output/总体QC报告.md] \\
        [--html-out ./qc_output/QC可视化报告.html] \\
        [--project-name <名字>]

默认输出到 <qc_dir>/总体QC报告.md 与 <qc_dir>/QC可视化报告.html。
"""
import argparse
import glob
import html as html_mod
import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from qc_lib import to_num  # noqa: E402

import markdown
import openpyxl


# ============================================================
# Part 1: markdown 总报告合成（原 merge_qc.py）
# ============================================================

LEVEL_ORDER = ["CRITICAL", "MAJOR", "MINOR", "SUGGESTION"]
LEVEL_CN = {"CRITICAL": "严重", "MAJOR": "主要", "MINOR": "次要", "SUGGESTION": "建议"}


def load_qc_jsons(qc_dir):
    """读 qc_dir 下所有 qc_*.json；返回 list[dict]。"""
    results = []
    for p in sorted(glob.glob(os.path.join(qc_dir, "qc_*.json"))):
        with open(p, encoding="utf-8") as f:
            results.append(json.load(f))
    return results


def extract_table_no(title):
    """从标题抽表号：'表 14.4.1 疗效分级' -> '14.4.1'。"""
    m = re.search(r"表\s*([\d.]+)", title or "")
    return m.group(1).rstrip(".") if m else None


def cross_table_rules(results, baseline):
    """跨表规则：R-029 表号唯一、R-021 同一分析集跨表分母一致。"""
    findings = []

    # R-029 表号唯一
    nos = [extract_table_no(r["meta"].get("title")) for r in results]
    for n, c in Counter(x for x in nos if x).items():
        if c > 1:
            dup_titles = [r["meta"]["title"] for r in results
                          if extract_table_no(r["meta"].get("title")) == n]
            findings.append({
                "rule": "R-029", "level": "MINOR", "where": f"表号 {n}",
                "expected": "全文档唯一", "found": f"{c} 张共用",
                "note": "；".join(dup_titles),
            })

    # R-021 同一分析集跨表分母一致
    by_set = defaultdict(list)
    for r in results:
        s = r["meta"].get("analysis_set")
        tot = (r["meta"].get("n_by_group") or {}).get("合计")
        if s and tot is not None:
            by_set[s].append((r["meta"]["title"], to_num(tot)))

    for s, rows in by_set.items():
        base = None
        if baseline and s in baseline:
            base = to_num((baseline[s] or {}).get("合计"))
        if base is None:
            base = max((t for _, t in rows if t is not None), default=None)
        if base is None:
            continue
        for title, tot in rows:
            if tot is not None and tot > base:
                findings.append({
                    "rule": "R-021", "level": "MAJOR",
                    "where": f"{title}·{s}",
                    "expected": f"≤{base:g}", "found": f"{tot:g}",
                    "note": "合计超过该分析集基准人数（人群划分表/同集最大值）",
                })
    return findings


def render_markdown_report(results, cross, baseline, source_hint):
    lines = []
    total = len(results)
    concl = Counter(r.get("conclusion", "PASS") for r in results)
    passed = concl.get("PASS", 0)
    n_issues = sum(len(r.get("issues", [])) for r in results) + len(cross)
    n_pending = sum(r.get("pending", 0) for r in results)
    types = Counter(r["meta"].get("table_type", "other") for r in results)

    lines.append("# 临床试验表格内部 QC 报告\n")
    lines.append("## 概要\n")
    lines.append(f"- 覆盖：{source_hint or '（未指定源文件）'}")
    lines.append(f"- 表格总数：**{total}** ｜ 通过：**{passed}** ｜ "
                 f"问题总数：**{n_issues}** ｜ 待人工：**{n_pending}**")
    concl_bits = " / ".join(f"{k} {concl[k]}" for k in
                            ["CRITICAL", "MAJOR", "MINOR", "SUGGESTION", "PASS"]
                            if concl.get(k))
    lines.append(f"- 结论分布：{concl_bits or '—'}")
    type_bits = " / ".join(f"{k} {v}张" for k, v in types.most_common())
    lines.append(f"- 表型分布：{type_bits}")
    lines.append(f"- 分析集基准：{'已加载 ' + '、'.join(baseline) if baseline else '未提供（基准类规则按 ≤ 跳过或转人工）'}\n")

    lines.append("## 按表格汇总\n")
    lines.append("| 编号 | 表型 | 标题 | 结论 | 问题数 | 待人工 |")
    lines.append("|------|------|------|------|--------|--------|")
    for r in sorted(results, key=lambda x: x["meta"].get("table_index", "")):
        m = r["meta"]
        lines.append(f"| {m.get('table_index','')} | {m.get('table_type','')} | "
                     f"{m.get('title','')} | {r.get('conclusion','PASS')} | "
                     f"{len(r.get('issues', []))} | {r.get('pending', 0)} |")
    lines.append("")

    buckets = defaultdict(list)
    for r in results:
        for it in r.get("issues", []):
            if it["level"] in LEVEL_ORDER:
                buckets[it["level"]].append((r["meta"]["title"], it))
    for it in cross:
        buckets[it["level"]].append(("【跨表】", it))

    lines.append("## 发现的问题\n")
    any_issue = False
    for lv in LEVEL_ORDER:
        if not buckets[lv]:
            continue
        any_issue = True
        lines.append(f"### [{LEVEL_CN[lv]}] {lv}（{len(buckets[lv])}）\n")
        for where_tbl, it in buckets[lv]:
            note = f" — {it['note']}" if it.get("note") else ""
            lines.append(f"- **{it['rule']}** {where_tbl}·{it['where']}："
                         f"期望 {it['expected']}，实际 {it['found']}{note}")
        lines.append("")
    if not any_issue:
        lines.append("未发现 CRITICAL/MAJOR/MINOR/SUGGESTION 级问题。\n")

    pend = [(r["meta"]["title"], it) for r in results
            for it in r.get("issues", []) if it["level"] == "待人工"]
    if pend:
        lines.append("## 待人工复核\n")
        for title, it in pend:
            note = f" — {it['note']}" if it.get("note") else ""
            lines.append(f"- **{it['rule']}** {title}·{it['where']}{note}")
        lines.append("")

    ok = [r["meta"]["title"] for r in results if r.get("conclusion") == "PASS"]
    if ok:
        lines.append("## 已核查通过\n")
        for t in ok:
            lines.append(f"- {t} ✓")
        lines.append("")

    return "\n".join(lines)


# ============================================================
# Part 2: HTML 可视化（原 build_report_viewer.py）
# ============================================================

SEVERITY_LABEL = {
    "major":   "🔴 Major",
    "manual":  "🟡 待人工",
    "pass":    "🟢 核查无误",
    "none":    "⚪ 未核查",
}
SEVERITY_ORDER = ["major", "manual", "pass", "none"]


def classify_severity(conclusion, pending):
    """Decide severity bucket for one table. Priority: major > manual > pass."""
    c = (conclusion or "").upper()
    if "MAJOR" in c or "CRITICAL" in c or "FAIL" in c:
        return "major"
    if pending and pending > 0:
        return "manual"
    if "待人工" in (conclusion or "") or "MANUAL" in c or "MINOR" in c:
        return "manual"
    if "PASS" in c or "OK" in c or "核查无误" in (conclusion or ""):
        return "pass"
    return "none"


_RE_QC_FILE = re.compile(r"^qc_(?:ext_)?(\d+)\.md$")
_RE_META = re.compile(r"^##META_(\w+):\s*(.+?)\s*$", re.MULTILINE)
_RE_CONCLUSION_LINE = re.compile(
    r"(?:核查)?结论[^|\n]*[|｜:：]\s*\*{0,2}([^*|｜\n（()]+)", re.IGNORECASE
)


def parse_report(md_path):
    """Extract conclusion, pending count, title, and render HTML for one report."""
    text = md_path.read_text(encoding="utf-8")

    # Sidecar JSON is the most reliable source when present
    json_path = md_path.with_suffix(".json")
    conclusion = ""
    pending = 0
    title = ""
    if json_path.exists():
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
            conclusion = str(data.get("conclusion", "") or "")
            pending = int(data.get("pending", 0) or 0)
            meta = data.get("meta", {}) or {}
            title = str(meta.get("title", "") or "")
        except (json.JSONDecodeError, ValueError):
            pass

    if not conclusion or not title:
        for m in _RE_META.finditer(text):
            key, val = m.group(1).upper(), m.group(2).strip()
            if key == "CONCLUSION" and not conclusion:
                conclusion = val
            elif key == "TABLE" and not title:
                title = val

    if not conclusion:
        head = "\n".join(text.splitlines()[:30])
        m = _RE_CONCLUSION_LINE.search(head)
        if m:
            conclusion = m.group(1).strip()

    if not title:
        h1 = re.search(r"^#\s+(.+?)\s*$", text, re.MULTILINE)
        if h1:
            title = h1.group(1).strip()

    severity = classify_severity(conclusion, pending)
    html = markdown.markdown(text, extensions=["tables", "fenced_code", "sane_lists"])

    return {
        "conclusion": conclusion or "未知",
        "pending": pending,
        "severity": severity,
        "title": title,
        "html": html,
    }


_SEV_RANK = {"major": 0, "manual": 1, "pass": 2, "none": 3}


def worst_severity(sevs):
    """Pick the worst (lowest rank) severity from a list; ignore falsy entries."""
    sevs = [s for s in sevs if s]
    return min(sevs, key=lambda s: _SEV_RANK[s]) if sevs else "none"


def load_md_reports(reports_dir):
    """Return {table_idx (int): {"internal": rep|None, "external": rep|None}}.

    Both `qc_<idx>.md` (Phase 4/5 内部一致性) and `qc_ext_<idx>.md` (Phase 3
    外部对照) are loaded under the same idx so `build_html_report` can present
    them together.
    """
    reports = {}
    for p in sorted(reports_dir.glob("qc_*.md")):
        m = _RE_QC_FILE.match(p.name)
        if not m:
            continue
        idx = int(m.group(1))
        slot = "external" if p.name.startswith("qc_ext_") else "internal"
        reports.setdefault(idx, {"internal": None, "external": None})[slot] = parse_report(p)
    return reports


def merge_reports(pair):
    """Combine internal + external reports for one table.

    - 两个都有：外部 section 在上，内部 section 在下，中间用虚线分隔；
      严重度取二者较差，pending 求和，结论合成 `外部:X / 内部:Y`。
    - 只有一份：单独一段，标签对应来源；结论、严重度、pending 沿用该份。
    - 两个都没有：返回 None，前端走"未核查"占位。
    """
    internal = pair.get("internal") if pair else None
    external = pair.get("external") if pair else None
    if not internal and not external:
        return None

    parts = []
    if external:
        parts.append(
            '<section class="qc-section">'
            '<h2 class="qc-section-title">外部对照核查</h2>'
            + external["html"] + '</section>'
        )
    if internal:
        parts.append(
            '<section class="qc-section">'
            '<h2 class="qc-section-title">内部一致性核查</h2>'
            + internal["html"] + '</section>'
        )

    sev = worst_severity([
        external["severity"] if external else None,
        internal["severity"] if internal else None,
    ])
    pending = (external["pending"] if external else 0) + (internal["pending"] if internal else 0)

    if external and internal:
        conclusion = f"外部:{external['conclusion']} / 内部:{internal['conclusion']}"
    elif external:
        conclusion = external["conclusion"]
    else:
        conclusion = internal["conclusion"]

    title = (external and external["title"]) or (internal and internal["title"]) or ""

    return {
        "severity": sev,
        "pending": pending,
        "html": "\n".join(parts),
        "conclusion": conclusion,
        "title": title,
    }


_RE_TABLE_FILE = re.compile(r"^(\d+)-(.+)\.xlsx$")
_KNOWN_ANALYSIS_SETS = {"FAS", "ITT", "mITT", "PPS", "SS", "随机化人群", "-"}


def xlsx_to_html(xlsx_path):
    """Render an xlsx as a single <table> (first sheet only, merged cells supported)."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=False)
    ws = wb.active

    span_map = {}
    skip = set()
    for mr in ws.merged_cells.ranges:
        rs = mr.max_row - mr.min_row + 1
        cs = mr.max_col - mr.min_col + 1
        span_map[(mr.min_row, mr.min_col)] = (rs, cs)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    skip.add((r, c))

    parts = ['<table class="xlsx-table">']
    for r in range(1, ws.max_row + 1):
        parts.append("<tr>")
        tag = "th" if r == 1 else "td"
        for c in range(1, ws.max_column + 1):
            if (r, c) in skip:
                continue
            val = ws.cell(r, c).value
            text = "" if val is None else str(val)
            cell_html = html_mod.escape(text).replace("\n", "<br>") or "&nbsp;"
            attrs = ""
            if (r, c) in span_map:
                rs, cs = span_map[(r, c)]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            parts.append(f"<{tag}{attrs}>{cell_html}</{tag}>")
        parts.append("</tr>")
    parts.append("</table>")
    return "\n".join(parts)


def load_xlsx_tables(tables_dir):
    """Return {table_idx (int): table dict} from a directory of NN-*.xlsx files."""
    tables = {}
    for p in sorted(tables_dir.glob("*.xlsx")):
        m = _RE_TABLE_FILE.match(p.name)
        if not m:
            continue
        idx = int(m.group(1))
        stem = m.group(2)
        type_hint = ""
        title_hint = stem
        if "-" in stem:
            front, tail = stem.rsplit("-", 1)
            if tail in _KNOWN_ANALYSIS_SETS and "-" in front:
                type_hint, title_hint = front.split("-", 1)
            else:
                parts = stem.split("-", 1)
                type_hint = parts[0]
                title_hint = parts[1] if len(parts) > 1 else stem
        tm = re.search(r"(表\s*\d+(?:\.\d+)*)\s*(.*)", title_hint)
        if tm:
            table_number = tm.group(1).strip()
            short_name = tm.group(2).strip().rstrip("-").strip()
        else:
            table_number = f"表 {idx}"
            short_name = title_hint
        tables[idx] = {
            "index": idx,
            "filename": p.name,
            "type": type_hint,
            "table_number": table_number,
            "short_name": short_name,
            "html": xlsx_to_html(p),
        }
    return tables


CSS = """
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html, body { height: 100%; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", Roboto, sans-serif; }
body { display: flex; flex-direction: column; overflow: hidden; background: #f8fafc; color: #0f172a; }

/* Header */
.top-bar {
  display: flex; align-items: center; gap: 16px;
  padding: 14px 24px; background: #fff;
  border-bottom: 1px solid #e2e8f0; flex-shrink: 0;
}
.top-bar h1 { font-size: 16px; font-weight: 700; }
.top-bar .subtitle { font-size: 12px; color: #64748b; margin-left: 4px; }
.top-bar .stats { display: flex; gap: 8px; margin-left: auto; font-size: 12px; flex-wrap: wrap; }
.top-bar .stats span { padding: 5px 12px; border-radius: 9999px; font-weight: 600; white-space: nowrap; }
.stat-total   { background: #e0e7ff; color: #3730a3; }
.stat-major   { background: #fee2e2; color: #991b1b; }
.stat-manual  { background: #fef9c3; color: #854d0e; }
.stat-pass    { background: #dcfce7; color: #166534; }
.stat-none    { background: #f1f5f9; color: #64748b; }

/* Main layout */
.main { display: flex; flex: 1; overflow: hidden; }

/* Left sidebar */
.sidebar {
  width: 400px; min-width: 320px; flex-shrink: 0;
  display: flex; flex-direction: column;
  border-right: 1px solid #e2e8f0; background: #fff;
}
.sidebar .search {
  padding: 14px 16px; border-bottom: 1px solid #f1f5f9;
}
.sidebar .search input {
  width: 100%; padding: 9px 14px;
  border: 1px solid #e2e8f0; border-radius: 10px;
  font-size: 13px; outline: none; background: #f8fafc;
  transition: border-color 0.15s, box-shadow 0.15s;
}
.sidebar .search input:focus {
  border-color: #2563eb; background: #fff;
  box-shadow: 0 0 0 3px rgba(37,99,235,0.12);
}
.sidebar .filter-chips {
  display: flex; gap: 6px; padding: 0 16px 12px; flex-wrap: wrap;
}
.sidebar .filter-chips button {
  font-size: 11px; padding: 4px 10px; border-radius: 9999px;
  border: 1px solid #e2e8f0; background: #f8fafc; color: #475569;
  cursor: pointer; font-weight: 600;
  transition: all 0.15s;
}
.sidebar .filter-chips button.active {
  background: #1e293b; color: #fff; border-color: #1e293b;
}
.sidebar .filter-chips button:hover:not(.active) { background: #eff6ff; }

.table-list { flex: 1; overflow-y: auto; padding: 4px 8px 12px; }
.tl-item {
  display: flex; align-items: center; gap: 10px;
  padding: 9px 12px; margin: 2px 0;
  cursor: pointer; border-radius: 10px;
  transition: background 0.1s;
  border: 1px solid transparent;
}
.tl-item:hover { background: #eff6ff; }
.tl-item.hovering { background: #fef3c7; border-color: #fde68a; }
.tl-item.active { background: #dbeafe; border-color: #93c5fd; }
.tl-item .tl-idx {
  font-size: 11px; color: #94a3b8; font-weight: 700;
  min-width: 24px; text-align: right;
}
.tl-item .tl-title {
  flex: 1; font-size: 13px; line-height: 1.35; color: #334155;
  overflow: hidden; text-overflow: ellipsis;
  display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical;
}
.tl-item .tl-badge {
  font-size: 11px; padding: 2px 8px; border-radius: 9999px;
  white-space: nowrap; font-weight: 600; flex-shrink: 0;
}
.badge-major  { background: #fee2e2; color: #991b1b; }
.badge-manual { background: #fef9c3; color: #854d0e; }
.badge-pass   { background: #dcfce7; color: #166534; }
.badge-none   { background: #f1f5f9; color: #94a3b8; }

.empty { padding: 24px; text-align: center; color: #94a3b8; font-size: 13px; }

/* Right panel */
.content { flex: 1; display: flex; flex-direction: column; overflow: hidden; }
.content-header {
  display: flex; align-items: center; gap: 12px;
  padding: 12px 24px; background: #fff;
  border-bottom: 1px solid #e2e8f0; flex-shrink: 0;
}
.content-header .title { font-size: 14px; font-weight: 700; color: #1e293b; flex: 1; }
.content-header .pinned { font-size: 11px; padding: 2px 8px; border-radius: 9999px; background: #dbeafe; color: #1e40af; font-weight: 600; }
.toggle { display: inline-flex; background: #f1f5f9; border-radius: 10px; padding: 3px; gap: 2px; }
.toggle button {
  padding: 6px 14px; border: none; border-radius: 8px;
  background: transparent; color: #64748b; cursor: pointer;
  font-size: 12px; font-weight: 600;
  transition: all 0.15s;
}
.toggle button.active { background: #fff; color: #1e293b; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
.toggle button:disabled { opacity: 0.4; cursor: not-allowed; }

.content-body { flex: 1; overflow-y: auto; padding: 24px; }

.placeholder {
  display: flex; flex-direction: column; align-items: center; justify-content: center;
  height: 100%; color: #94a3b8; text-align: center;
}
.placeholder .icon { font-size: 56px; opacity: 0.4; margin-bottom: 14px; }
.placeholder p { font-size: 14px; line-height: 1.7; max-width: 380px; }
.placeholder .tip { font-size: 12px; margin-top: 8px; color: #cbd5e1; }

.report-shell { max-width: 1080px; }
.report-shell h1, .report-shell h2, .report-shell h3 { margin: 16px 0 10px; color: #0f172a; }
.report-shell h1 { font-size: 19px; }
.report-shell h2 { font-size: 16px; padding-bottom: 6px; border-bottom: 1px solid #e2e8f0; }
.report-shell h3 { font-size: 14px; }
.report-shell p { margin: 8px 0; line-height: 1.7; font-size: 13px; color: #334155; }
.report-shell ul, .report-shell ol { margin: 8px 0; padding-left: 24px; }
.report-shell li { line-height: 1.7; font-size: 13px; color: #334155; }
.report-shell blockquote {
  border-left: 3px solid #2563eb; padding: 8px 14px;
  margin: 10px 0; background: #eff6ff;
  font-size: 13px; border-radius: 0 8px 8px 0; color: #1e3a8a;
}
.report-shell code { background: #f1f5f9; padding: 2px 6px; border-radius: 4px; font-size: 12px; color: #0f172a; }
.report-shell pre { background: #0f172a; color: #e2e8f0; padding: 14px; border-radius: 10px; overflow-x: auto; font-size: 12px; margin: 12px 0; }
.report-shell hr { border: none; border-top: 1px solid #e2e8f0; margin: 18px 0; }
.report-shell a { color: #2563eb; text-decoration: none; }
.report-shell a:hover { text-decoration: underline; }

.report-shell table {
  border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 12px;
  background: #fff; box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}
.report-shell th, .report-shell td {
  border: 1px solid #e2e8f0; padding: 6px 10px;
  text-align: left; vertical-align: top;
}
.report-shell th { background: #f8fafc; font-weight: 600; color: #475569; }
.report-shell tr:nth-child(even) td { background: #fafafa; }

/* Strip META lines that aren't useful in rendered view */
.report-shell h2:first-child + p:has(+ h2),
.report-shell p:has(>code:only-child) { font-family: ui-monospace, monospace; font-size: 11px; color: #94a3b8; }

/* Original xlsx table */
.xlsx-table {
  border-collapse: collapse; width: 100%; font-size: 12px;
  background: #fff; box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}
.xlsx-table th, .xlsx-table td {
  border: 1px solid #e2e8f0; padding: 6px 10px;
  text-align: left; vertical-align: top;
}
.xlsx-table th { background: #dbeafe; color: #1e3a8a; font-weight: 600; }
.xlsx-table tr:nth-child(even) td { background: #f8fafc; }
.xlsx-meta {
  padding: 12px 16px; background: #fff;
  border: 1px solid #e2e8f0; border-radius: 10px;
  margin-bottom: 14px; font-size: 12px; color: #64748b;
}
.xlsx-meta strong { color: #1e293b; }

/* 核查类型标签（外部对照核查 / 内部一致性核查） */
.qc-section + .qc-section {
  margin-top: 28px; padding-top: 18px;
  border-top: 2px dashed #cbd5e1;
}
.qc-section-title {
  font-size: 15px !important; color: #2563eb !important;
  border-bottom: none !important; padding-bottom: 0 !important;
  margin: 0 0 14px !important;
}
.qc-section-title::before {
  content: ""; display: inline-block;
  width: 3px; height: 14px; background: #2563eb;
  margin-right: 8px; vertical-align: -2px; border-radius: 2px;
}

/* Scrollbar */
::-webkit-scrollbar { width: 8px; height: 8px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }
"""


def build_html(items, stats, project_name):
    """Generate the full self-contained HTML page."""
    items_json = json.dumps(items, ensure_ascii=False)

    stat_parts = [f'<span class="stat-total">总计: {stats["total"]}</span>']
    for sev in ("major", "manual", "pass", "none"):
        n = stats.get(sev, 0)
        if n:
            label = SEVERITY_LABEL[sev].split(" ", 1)[-1]
            stat_parts.append(f'<span class="stat-{sev}">{label}: {n}</span>')
    stats_html = "\n    ".join(stat_parts)

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{html_mod.escape(project_name)} — 表格审查报告</title>
<style>{CSS}</style>
</head>
<body>

<header class="top-bar">
  <h1>📋 表格审查报告</h1>
  <span class="subtitle">{html_mod.escape(project_name)}</span>
  <div class="stats">
    {stats_html}
  </div>
</header>

<div class="main">

  <aside class="sidebar">
    <div class="search">
      <input id="searchInput" type="text" placeholder="按表号或表名搜索…">
    </div>
    <div class="filter-chips" id="filterChips">
      <button data-filter="all" class="active">全部</button>
      <button data-filter="major">🔴 Major</button>
      <button data-filter="manual">🟡 待人工</button>
      <button data-filter="pass">🟢 核查无误</button>
      <button data-filter="none">⚪ 未核查</button>
    </div>
    <div id="tableList" class="table-list"></div>
  </aside>

  <section class="content">
    <div class="content-header">
      <div class="title" id="contentTitle">选择左侧表格查看审查结果</div>
      <span class="pinned" id="pinnedFlag" style="display:none;">📌 已固定</span>
      <div class="toggle" id="viewToggle">
        <button data-view="report" class="active">审查意见</button>
        <button data-view="table">原始表格</button>
      </div>
    </div>
    <div class="content-body" id="contentBody">
      <div class="placeholder">
        <div class="icon">📊</div>
        <p>悬停左侧表格即可预览审查报告</p>
        <p class="tip">点击表格固定显示 · 顶部切换"审查意见 / 原始表格" · Esc 取消固定</p>
      </div>
    </div>
  </section>

</div>

<script>
const ITEMS = {items_json};
const SEV_LABEL = {{
  major: '🔴 Major',
  manual: '🟡 待人工',
  pass: '🟢 核查无误',
  none: '⚪ 未核查',
}};

const tableList   = document.getElementById('tableList');
const contentBody = document.getElementById('contentBody');
const contentTitle= document.getElementById('contentTitle');
const pinnedFlag  = document.getElementById('pinnedFlag');
const searchInput = document.getElementById('searchInput');

let activeIdx = null;       // pinned index
let hoverIdx  = null;
let currentView = 'report'; // 'report' | 'table'
let currentFilter = 'all';

function render(filter, q) {{
  q = (q || '').toLowerCase().trim();
  tableList.innerHTML = '';
  const rows = ITEMS.filter(it => {{
    if (filter && filter !== 'all' && it.severity !== filter) return false;
    if (!q) return true;
    return (it.table_number || '').toLowerCase().includes(q)
        || (it.short_name || '').toLowerCase().includes(q)
        || String(it.index).includes(q);
  }});
  if (rows.length === 0) {{
    tableList.innerHTML = '<div class="empty">没有匹配的表格</div>';
    return;
  }}
  rows.forEach(it => {{
    const div = document.createElement('div');
    div.className = 'tl-item';
    if (it.index === activeIdx) div.classList.add('active');
    div.dataset.idx = it.index;
    div.innerHTML =
      '<span class="tl-idx">#' + String(it.index).padStart(2, '0') + '</span>' +
      '<span class="tl-title">' + escapeHtml(it.table_number) + ' ' + escapeHtml(it.short_name) + '</span>' +
      '<span class="tl-badge badge-' + it.severity + '">' + SEV_LABEL[it.severity] + '</span>';
    div.addEventListener('mouseenter', () => onHover(it.index));
    div.addEventListener('mouseleave', onLeave);
    div.addEventListener('click', () => onClick(it.index));
    tableList.appendChild(div);
  }});
}}

function onHover(idx) {{
  hoverIdx = idx;
  document.querySelectorAll('.tl-item').forEach(el => {{
    el.classList.toggle('hovering', parseInt(el.dataset.idx) === idx && idx !== activeIdx);
  }});
  showItem(idx, /*ephemeral*/ activeIdx === null || activeIdx !== idx);
}}

function onLeave() {{
  hoverIdx = null;
  document.querySelectorAll('.tl-item').forEach(el => el.classList.remove('hovering'));
  if (activeIdx !== null) showItem(activeIdx, false);
  else showPlaceholder();
}}

function onClick(idx) {{
  if (activeIdx === idx) {{
    activeIdx = null;
    showPlaceholder();
  }} else {{
    activeIdx = idx;
    showItem(idx, false);
  }}
  render(currentFilter, searchInput.value);
}}

function showItem(idx, ephemeral) {{
  const it = ITEMS.find(x => x.index === idx);
  if (!it) return;
  contentTitle.textContent = it.table_number + ' ' + it.short_name;
  pinnedFlag.style.display = (activeIdx === idx && !ephemeral) ? '' : 'none';

  const tableBtn = document.querySelector('#viewToggle button[data-view="table"]');
  tableBtn.disabled = !it.table_html;
  const reportBtn = document.querySelector('#viewToggle button[data-view="report"]');
  reportBtn.disabled = !it.report_html;

  let view = currentView;
  if (view === 'table' && !it.table_html) view = 'report';
  if (view === 'report' && !it.report_html) view = 'table';

  if (view === 'report' && it.report_html) {{
    contentBody.innerHTML = '<div class="report-shell">' + it.report_html + '</div>';
  }} else if (view === 'table' && it.table_html) {{
    const meta = '<div class="xlsx-meta">' +
      '<strong>' + escapeHtml(it.table_number) + '</strong> ' + escapeHtml(it.short_name) +
      ' · 文件: <code>' + escapeHtml(it.filename || '') + '</code>' +
      '</div>';
    contentBody.innerHTML = meta + it.table_html;
  }} else {{
    contentBody.innerHTML = '<div class="placeholder"><div class="icon">📭</div>' +
      '<p>该表格' + (view === 'report' ? '没有审查报告' : '没有原始表格文件') + '</p></div>';
  }}
}}

function showPlaceholder() {{
  contentTitle.textContent = '选择左侧表格查看审查结果';
  pinnedFlag.style.display = 'none';
  contentBody.innerHTML =
    '<div class="placeholder"><div class="icon">📊</div>' +
    '<p>悬停左侧表格即可预览审查报告</p>' +
    '<p class="tip">点击表格固定显示 · 顶部切换"审查意见 / 原始表格" · Esc 取消固定</p></div>';
}}

document.getElementById('viewToggle').addEventListener('click', e => {{
  if (e.target.tagName !== 'BUTTON' || e.target.disabled) return;
  currentView = e.target.dataset.view;
  document.querySelectorAll('#viewToggle button').forEach(b =>
    b.classList.toggle('active', b.dataset.view === currentView));
  const target = activeIdx !== null ? activeIdx : hoverIdx;
  if (target !== null) showItem(target, activeIdx !== target);
}});

document.getElementById('filterChips').addEventListener('click', e => {{
  if (e.target.tagName !== 'BUTTON') return;
  currentFilter = e.target.dataset.filter;
  document.querySelectorAll('#filterChips button').forEach(b =>
    b.classList.toggle('active', b.dataset.filter === currentFilter));
  render(currentFilter, searchInput.value);
}});

searchInput.addEventListener('input', () => render(currentFilter, searchInput.value));

document.addEventListener('keydown', e => {{
  if (e.key === 'Escape' && activeIdx !== null) {{
    activeIdx = null;
    showPlaceholder();
    render(currentFilter, searchInput.value);
  }}
}});

function escapeHtml(s) {{
  const d = document.createElement('div');
  d.textContent = String(s == null ? '' : s);
  return d.innerHTML;
}}

render('all', '');
</script>

</body>
</html>"""


def build_html_report(qc_dir, tables_dir, output_path, project_name):
    """从 qc_dir 里的 qc_*.md + tables_dir 里的 xlsx 组装 HTML 报告。"""
    reports_dir = Path(qc_dir).expanduser().resolve()
    tables_dir_p = Path(tables_dir).expanduser().resolve()

    print(f"[HTML] 读取审查报告: {reports_dir}")
    reports = load_md_reports(reports_dir)
    print(f"       → {len(reports)} 份")

    print(f"[HTML] 读取原始表格: {tables_dir_p}")
    tables = load_xlsx_tables(tables_dir_p)
    print(f"       → {len(tables)} 张")

    all_idx = sorted(set(reports.keys()) | set(tables.keys()))
    items = []
    stats = {"total": 0, "major": 0, "manual": 0, "pass": 0, "none": 0}
    for idx in all_idx:
        merged = merge_reports(reports.get(idx))
        tab = tables.get(idx)
        if tab:
            table_number = tab["table_number"]
            short_name = tab["short_name"]
        elif merged and merged["title"]:
            tm = re.search(r"(表\s*\d+(?:\.\d+)*)\s*(.*)", merged["title"])
            if tm:
                table_number = tm.group(1).strip()
                short_name = tm.group(2).strip()
            else:
                table_number = f"表 {idx}"
                short_name = merged["title"]
        else:
            table_number = f"表 {idx}"
            short_name = "(无标题)"

        severity = merged["severity"] if merged else "none"
        stats[severity] = stats.get(severity, 0) + 1
        stats["total"] += 1

        items.append({
            "index": idx,
            "table_number": table_number,
            "short_name": short_name,
            "filename": tab["filename"] if tab else None,
            "severity": severity,
            "conclusion": merged["conclusion"] if merged else "未核查",
            "pending": merged["pending"] if merged else 0,
            "report_html": merged["html"] if merged else None,
            "table_html": tab["html"] if tab else None,
        })

    print(f"[HTML] 统计: major={stats['major']}, 待人工={stats['manual']}, "
          f"核查无误={stats['pass']}, 未核查={stats['none']}")

    html = build_html(items, stats, project_name)
    output_path.write_text(html, encoding="utf-8")
    size_kb = output_path.stat().st_size / 1024
    print(f"[HTML] ✅ {output_path} ({size_kb:.0f} KB)")


# ============================================================
# 统一 CLI
# ============================================================

def main():
    ap = argparse.ArgumentParser(
        description="Phase 6 一站式：markdown 总报告 + HTML 可视化报告"
    )
    ap.add_argument("qc_dir", help="存放 qc_*.json / qc_*.md 的目录")
    ap.add_argument("--tables-dir", required=True,
                    help="存放原始 xlsx 的目录（HTML 报告用于内嵌原表）")
    ap.add_argument("--baseline", default=None,
                    help="baseline.json 路径；用于 R-021 跨表核查（可选）")
    ap.add_argument("--source", default=None,
                    help="源文件名（写进 markdown 概要）")
    ap.add_argument("--md-out", default=None,
                    help="markdown 输出路径（默认 <qc_dir>/总体QC报告.md）")
    ap.add_argument("--html-out", default=None,
                    help="HTML 输出路径（默认 <qc_dir>/QC可视化报告.html）")
    ap.add_argument("--project-name", default=None,
                    help="HTML 页面标题里的项目名（默认取 qc_dir 上一级目录名）")
    args = ap.parse_args()

    qc_dir = args.qc_dir
    if not os.path.isdir(qc_dir):
        sys.exit(f"错误：qc_dir 不存在 {qc_dir}")
    if not os.path.isdir(args.tables_dir):
        sys.exit(f"错误：tables-dir 不存在 {args.tables_dir}")

    # ---- Part 1: markdown 总报告 ----
    results = load_qc_jsons(qc_dir)
    if not results:
        sys.exit(f"错误：{qc_dir} 下没有 qc_*.json")

    baseline = {}
    if args.baseline and os.path.exists(args.baseline):
        with open(args.baseline, encoding="utf-8") as f:
            baseline = json.load(f)

    cross = cross_table_rules(results, baseline)
    md_body = render_markdown_report(results, cross, baseline, args.source)

    md_out = args.md_out or os.path.join(qc_dir, "总体QC报告.md")
    os.makedirs(os.path.dirname(os.path.abspath(md_out)) or ".", exist_ok=True)
    with open(md_out, "w", encoding="utf-8") as f:
        f.write(md_body)
    print(f"[markdown] 合并 {len(results)} 张表 · 跨表问题 {len(cross)} 条 → {md_out}")

    # ---- Part 2: HTML 可视化报告 ----
    html_out = args.html_out or os.path.join(qc_dir, "QC可视化报告.html")
    output_path = Path(html_out).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    project_name = args.project_name or Path(qc_dir).resolve().parent.name

    build_html_report(qc_dir, args.tables_dir, output_path, project_name)


if __name__ == "__main__":
    main()
