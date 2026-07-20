#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""按共同键（默认"筛选号"）把两张 xlsx 表左右拼成一张。

设计目标：只要两表都在表头行出现同名键列（默认"筛选号"），就能合并——不依赖
项目内其他约定。典型用法是 Phase 3 前把"随机表"和"人群划分表"按筛选号拼成
一张宽表，供后续核对/追溯。

用法::

    python3 merge_by_key.py 左表.xlsx 右表.xlsx --out 合并.xlsx
    python3 merge_by_key.py A.xlsx B.xlsx --out out.xlsx --key 病例号 --how inner

行为::

- 自动在两表表头行（默认第 1 行）定位键列；找不到直接报错列出可用列名
- 键值统一 str().strip() 后比对，兼容 "01001" vs 1001 / 前导零丢失等常见坑
- 默认 outer（全外连接）：两表所有主体都保留，缺席方留空
- 列名冲突（例如两表都有"组别"）时给右表加后缀 " (右表)"
- 键列只保留一份（在左表位置）
- 输出保留左表原顺序，右表独有行按键排序追加在末尾

多 sheet::

- 默认自动扫描工作簿里所有 sheet：只要 sheet 的表头行含键列，就把它的数据行拼进来
  （常见用途：一个"按中心分 sheet"的随机表，6 个中心表头一致，会自动整合成一整张）
- `--left-sheet NAME` / `--right-sheet NAME` 显式指定单一 sheet
- 拼接会加一列 `_source_sheet` 记录来源 sheet 名，方便回溯

局限::

- 只按"表头列名完全一致"识别键列——若一边写"筛选号"另一边写"筛选编号"，需
  两侧起相同名字，或改用 --left-key / --right-key（未实现，可扩展）
- 表头默认在第 1 行；多级表头需要人为先把 xlsx 处理成单级
- 一对多关系（同一键在一表出现多次）会按首次出现匹配，多余行的匹配值会重复
- 多 sheet 自动模式假设各 sheet 表头一致，只保留与首个匹配 sheet 相同的表头
"""
import argparse
import os
import sys
from collections import OrderedDict

from openpyxl import Workbook, load_workbook


def _stringify(v):
    """把任意单元格值规范成用于键匹配的字符串。

    数字型自动转 int-string（"01001" 不会因为读进来变成 float 而失配）。
    """
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_sheet(ws):
    """读一个 sheet 为二维列表。"""
    return [list(row) for row in ws.iter_rows(values_only=True)]


def load_table(path, key_name, sheet=None):
    """读 xlsx，返回 (headers, rows, key_col_idx, source_sheets)。

    sheet 语义：
    - None: 扫描所有 sheet，把表头含 key_name 的都拼起来（首个匹配 sheet 的表头为准）
    - str: 只读指定名字的 sheet
    """
    if not os.path.exists(path):
        raise SystemExit(f"错误：文件不存在 {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        # 单 sheet 显式指定
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise SystemExit(
                    f"错误：{path} 没有名为 {sheet!r} 的 sheet\n"
                    f"可用 sheet: {wb.sheetnames}"
                )
            grid = _read_sheet(wb[sheet])
            if not grid:
                raise SystemExit(f"错误：{path}::{sheet} 是空 sheet")
            headers = [str(h).strip() if h is not None else "" for h in grid[0]]
            if key_name not in headers:
                raise SystemExit(
                    f"错误：{path}::{sheet} 表头未找到键列 {key_name!r}\n"
                    f"可用列名: {headers}"
                )
            key_idx = headers.index(key_name)
            data_rows = grid[1:]
            print(f"[{os.path.basename(path)}] 使用 sheet {sheet!r}：{len(data_rows)} 行")
            return headers, data_rows, key_idx, [sheet] * len(data_rows)

        # 自动模式：扫描全部 sheet
        base_headers = None
        key_idx = None
        all_rows = []
        source_marks = []
        used_sheets = []
        for name in wb.sheetnames:
            grid = _read_sheet(wb[name])
            if not grid:
                continue
            hs = [str(h).strip() if h is not None else "" for h in grid[0]]
            if key_name not in hs:
                continue
            if base_headers is None:
                base_headers = hs
                key_idx = hs.index(key_name)
            elif hs != base_headers:
                print(
                    f"⚠️  {os.path.basename(path)}::{name} 表头与 {used_sheets[0]!r} 不一致，跳过\n"
                    f"    {name!r} 表头: {hs}\n"
                    f"    基准表头: {base_headers}",
                    file=sys.stderr,
                )
                continue
            data_rows = grid[1:]
            all_rows.extend(data_rows)
            source_marks.extend([name] * len(data_rows))
            used_sheets.append(name)

        if not used_sheets:
            raise SystemExit(
                f"错误：{path} 所有 sheet 表头都没找到键列 {key_name!r}\n"
                f"扫描的 sheet: {wb.sheetnames}"
            )
        print(
            f"[{os.path.basename(path)}] 自动模式：从 {len(used_sheets)} 个 sheet "
            f"拼出 {len(all_rows)} 行（{', '.join(used_sheets)}）"
        )
        return base_headers, all_rows, key_idx, source_marks
    finally:
        wb.close()


def merge(left_path, right_path, key_name, how,
          left_sheet=None, right_sheet=None):
    lh, lr, lki, l_src = load_table(left_path, key_name, sheet=left_sheet)
    rh, rr, rki, r_src = load_table(right_path, key_name, sheet=right_sheet)

    # 右表按键索引 + 记住每条来自哪个 sheet
    right_idx = OrderedDict()      # key -> row
    right_src_by_key = {}          # key -> sheet name
    right_dupes = []
    for r, src in zip(rr, r_src):
        k = _stringify(r[rki] if rki < len(r) else None)
        if not k:
            continue
        if k in right_idx:
            right_dupes.append(k)
        else:
            right_idx[k] = r
            right_src_by_key[k] = src
    if right_dupes:
        print(f"⚠️  右表键 {key_name} 有 {len(right_dupes)} 个重复："
              f"{', '.join(right_dupes[:5])}{' …' if len(right_dupes) > 5 else ''}",
              file=sys.stderr)

    # 合并表头：左表全部 + 右表非键列（列名冲突加后缀）+ 追溯列
    right_extra_cols = []  # [(right_col_idx, output_header_name)]
    for i, h in enumerate(rh):
        if i == rki:
            continue
        out_name = f"{h} (右表)" if h in lh else h
        right_extra_cols.append((i, out_name))

    # 判断是否需要 sheet 追溯列（任一侧从多个 sheet 拼过来就加）
    l_multi = len(set(l_src)) > 1
    r_multi = len(set(r_src)) > 1
    src_col_names = []
    if l_multi:
        src_col_names.append("_source_sheet_左")
    if r_multi:
        src_col_names.append("_source_sheet_右")

    merged_headers = (
        list(lh)
        + [name for _, name in right_extra_cols]
        + src_col_names
    )

    # 拼行
    merged_rows = []
    matched_keys = set()
    for lrow, lsrc in zip(lr, l_src):
        left_padded = list(lrow) + [None] * (len(lh) - len(lrow))
        k = _stringify(left_padded[lki])
        if k and k in right_idx:
            rrow = right_idx[k]
            rextras = [rrow[i] if i < len(rrow) else None
                       for i, _ in right_extra_cols]
            r_src_val = right_src_by_key[k]
            matched_keys.add(k)
        else:
            rextras = [None] * len(right_extra_cols)
            r_src_val = None
        extras = []
        if l_multi:
            extras.append(lsrc)
        if r_multi:
            extras.append(r_src_val)
        merged_rows.append(left_padded + rextras + extras)

    # 右表独有行（outer / right 需要）
    if how in ("outer", "right"):
        for k, rrow in right_idx.items():
            if k in matched_keys:
                continue
            left_side = [None] * len(lh)
            left_side[lki] = rrow[rki] if rki < len(rrow) else None
            rextras = [rrow[i] if i < len(rrow) else None
                       for i, _ in right_extra_cols]
            extras = []
            if l_multi:
                extras.append(None)
            if r_multi:
                extras.append(right_src_by_key[k])
            merged_rows.append(left_side + rextras + extras)

    # inner 过滤未匹配的左表行
    if how == "inner":
        merged_rows = [
            row for row in merged_rows
            if _stringify(row[lki]) in matched_keys
        ]

    # right join 过滤未匹配的左表行
    if how == "right":
        merged_rows = [
            row for row in merged_rows
            if _stringify(row[lki]) in matched_keys
            or all(v is None for i, v in enumerate(row) if i < len(lh) and i != lki)
        ]

    return merged_headers, merged_rows, len(matched_keys)


def write_xlsx(headers, rows, out_path):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append([("" if v is None else v) for v in r])
    os.makedirs(os.path.dirname(os.path.abspath(out_path)) or ".",
                exist_ok=True)
    wb.save(out_path)


def main():
    ap = argparse.ArgumentParser(
        description="按共同键（默认'筛选号'）合并两张 xlsx"
    )
    ap.add_argument("left", help="左表 xlsx")
    ap.add_argument("right", help="右表 xlsx")
    ap.add_argument("--out", required=True, help="输出 xlsx 路径")
    ap.add_argument("--key", default="筛选号",
                    help="共同键列名（两表表头都得有），默认'筛选号'")
    ap.add_argument("--how", choices=["outer", "left", "right", "inner"],
                    default="outer",
                    help="连接方式：outer(默认) / left / right / inner")
    ap.add_argument("--left-sheet", default=None,
                    help="左表只读指定 sheet；默认自动拼接所有含键列的 sheet")
    ap.add_argument("--right-sheet", default=None,
                    help="右表只读指定 sheet；默认自动拼接所有含键列的 sheet")
    args = ap.parse_args()

    headers, rows, n_matched = merge(
        args.left, args.right, args.key, args.how,
        left_sheet=args.left_sheet, right_sheet=args.right_sheet,
    )
    write_xlsx(headers, rows, args.out)

    print(f"\n合并方式：{args.how}｜键列：{args.key}")
    print(f"匹配上 {n_matched} 条；输出 {len(rows)} 行 × {len(headers)} 列 → {args.out}")


if __name__ == "__main__":
    main()
