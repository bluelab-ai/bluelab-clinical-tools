#!/usr/bin/env python3
"""
Build an interactive QC report viewer HTML page.

Input:
  - A directory of subagent-generated markdown reports (qc_NN.md / qc_NN.json)
  - A directory of original tables (xlsx, named NN-...xlsx)

Output:
  A single self-contained HTML file with:
    - Top stats capsules (major / 待人工 / 核查无误 / 未被核查)
    - Left sidebar: searchable table list with colored severity badges
    - Right panel: hover-to-preview QC report, click-to-pin, toggle between
      "审查意见" (rendered md report) and "原始表格" (rendered xlsx)

Usage:
    python3 build_report_viewer.py \\
        --reports-dir /path/to/qc_output \\
        --tables-dir  /path/to/tables_output \\
        --output      report-viewer.html
"""

import argparse
import json
import re
import html as html_mod
from pathlib import Path

import markdown
import openpyxl

try:
    from PIL import Image
    import base64 as b64_mod
    import io
    _HAS_PIL = True
except ImportError:
    _HAS_PIL = False


# ── Severity classification ─────────────────────────────────────────────────
SEVERITY_LABEL = {
    "major":   "🔴 Major",
    "manual":  "🟡 待人工",
    "pass":    "🟢 核查无误",
    "none":    "⚪ 未核查",
}
SEVERITY_ORDER = ["major", "manual", "pass", "none"]


def classify(conclusion: str, pending: int) -> str:
    """Decide severity bucket for one table. Priority: major > manual > pass."""
    c = (conclusion or "").upper()
    if "MAJOR" in c or "CRITICAL" in c or "FAIL" in c:
        return "major"
    if pending and pending > 0:
        return "manual"
    if "待人工" in (conclusion or "") or "MANUAL" in c:
        return "manual"
    if "PASS" in c or "OK" in c or "核查无误" in (conclusion or ""):
        return "pass"
    return "none"


# ── Reports parsing ────────────────────────────────────────────────────────
_RE_QC_FILE = re.compile(r"^qc_(?:ext_)?(\d+)\.md$")
_RE_META = re.compile(r"^##META_(\w+):\s*(.+?)\s*$", re.MULTILINE)
_RE_CONCLUSION_LINE = re.compile(
    r"(?:核查)?结论[^|\n]*[|｜:：]\s*\*{0,2}([^*|｜\n（()]+)", re.IGNORECASE
)


def parse_report(md_path: Path) -> dict:
    """Extract conclusion, pending count, title, and render HTML for one report."""
    text = md_path.read_text(encoding="utf-8")

    # Sidecar JSON is the most reliable source when present
    json_path = md_path.with_suffix(".json")
    conclusion = ""
    pending = 0
    title = ""
    if json_path.exists():
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
            conclusion = str(data.get("conclusion", "") or "")
            pending = int(data.get("pending", 0) or 0)
            meta = data.get("meta", {}) or {}
            title = str(meta.get("title", "") or "")
        except (json.JSONDecodeError, ValueError):
            pass

    # Fallback: parse meta lines from md
    if not conclusion or not title:
        for m in _RE_META.finditer(text):
            key, val = m.group(1).upper(), m.group(2).strip()
            if key == "CONCLUSION" and not conclusion:
                conclusion = val
            elif key == "TABLE" and not title:
                title = val

    # Fallback: look for "结论 | XXX" style row in the first 30 lines
    if not conclusion:
        head = "\n".join(text.splitlines()[:30])
        m = _RE_CONCLUSION_LINE.search(head)
        if m:
            conclusion = m.group(1).strip()

    # Last resort: derive title from H1
    if not title:
        h1 = re.search(r"^#\s+(.+?)\s*$", text, re.MULTILINE)
        if h1:
            title = h1.group(1).strip()

    severity = classify(conclusion, pending)

    html = markdown.markdown(
        text, extensions=["tables", "fenced_code", "sane_lists"]
    )

    return {
        "conclusion": conclusion or "未知",
        "pending": pending,
        "severity": severity,
        "title": title,
        "html": html,
    }


def load_reports(reports_dir: Path) -> dict[int, dict]:
    """Return {table_idx (int): report dict} from a directory of qc_NN.md files."""
    reports: dict[int, dict] = {}
    for p in sorted(reports_dir.glob("qc_*.md")):
        m = _RE_QC_FILE.match(p.name)
        if not m:
            continue
        idx = int(m.group(1))
        reports[idx] = parse_report(p)
    return reports


# ── Table parsing ──────────────────────────────────────────────────────────
_RE_TABLE_FILE = re.compile(r"^(\d+)-(.+)\.xlsx$")
_KNOWN_ANALYSIS_SETS = {"FAS", "ITT", "mITT", "PPS", "SS", "随机化人群", "-"}


def xlsx_to_html(xlsx_path: Path) -> str:
    """Render an xlsx as a single <table> (first sheet only, merged cells supported)."""
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True, read_only=False)
    ws = wb.active

    # Map merged cells: {(row, col): (rowspan, colspan)} for the anchor cell,
    # plus a set of cells to skip (those covered by an anchor).
    span_map: dict[tuple[int, int], tuple[int, int]] = {}
    skip: set[tuple[int, int]] = set()
    for mr in ws.merged_cells.ranges:
        rs = mr.max_row - mr.min_row + 1
        cs = mr.max_col - mr.min_col + 1
        span_map[(mr.min_row, mr.min_col)] = (rs, cs)
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                if (r, c) != (mr.min_row, mr.min_col):
                    skip.add((r, c))

    parts = ['<table class="xlsx-table">']
    for r in range(1, ws.max_row + 1):
        parts.append("<tr>")
        tag = "th" if r == 1 else "td"
        for c in range(1, ws.max_column + 1):
            if (r, c) in skip:
                continue
            val = ws.cell(r, c).value
            text = "" if val is None else str(val)
            cell_html = html_mod.escape(text).replace("\n", "<br>") or "&nbsp;"
            attrs = ""
            if (r, c) in span_map:
                rs, cs = span_map[(r, c)]
                if rs > 1:
                    attrs += f' rowspan="{rs}"'
                if cs > 1:
                    attrs += f' colspan="{cs}"'
            parts.append(f"<{tag}{attrs}>{cell_html}</{tag}>")
        parts.append("</tr>")
    parts.append("</table>")
    return "\n".join(parts)


def load_tables(tables_dir: Path) -> dict[int, dict]:
    """Return {table_idx (int): table dict} from a directory of NN-*.xlsx files."""
    tables: dict[int, dict] = {}
    for p in sorted(tables_dir.glob("*.xlsx")):
        m = _RE_TABLE_FILE.match(p.name)
        if not m:
            continue
        idx = int(m.group(1))
        stem = m.group(2)
        # filename: "<type>-<title>-<analysis_set>.xlsx" or "<type>-<title>.xlsx"
        # strip known analysis set suffix first for clean title extraction
        type_hint = ""
        title_hint = stem
        if "-" in stem:
            front, tail = stem.rsplit("-", 1)
            if tail in _KNOWN_ANALYSIS_SETS and "-" in front:
                type_hint, title_hint = front.split("-", 1)
            else:
                parts = stem.split("-", 1)
                type_hint = parts[0]
                title_hint = parts[1] if len(parts) > 1 else stem
        tm = re.search(r"(表\s*\d+(?:\.\d+)*)\s*(.*)", title_hint)
        if tm:
            table_number = tm.group(1).strip()
            short_name = tm.group(2).strip().rstrip("-").strip()
        else:
            table_number = f"表 {idx}"
            short_name = title_hint
        tables[idx] = {
            "index": idx,
            "filename": p.name,
            "type": type_hint,
            "table_number": table_number,
            "short_name": short_name,
            "html": xlsx_to_html(p),
        }
    return tables


# ── HTML builder ────────────────────────────────────────────────────────────
CSS = """
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
html, body { height: 100%; font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", Roboto, sans-serif; }
body { display: flex; flex-direction: column; overflow: hidden; background: #f8fafc; color: #0f172a; }

/* Header */
.top-bar {
  display: flex; align-items: center; gap: 16px;
  padding: 16px 24px; background: #fff;
  border-bottom: 1px solid rgba(226,232,240,0.7); flex-shrink: 0;
}
.top-bar h1 { font-size: 16px; font-weight: 700; color: #0f172a; }
.top-bar .subtitle { font-size: 12px; color: #64748b; margin-left: 12px; }
.top-bar .stats { display: flex; gap: 10px; margin-left: auto; font-size: 12px; flex-wrap: wrap; }
.top-bar .stats span { padding: 4px 12px; border-radius: 9999px; font-weight: 600; white-space: nowrap; }
.stat-total   { background: #e0e7ff; color: #3730a3; }
.stat-major   { background: #fee2e2; color: #991b1b; }
.stat-manual  { background: #fef9c3; color: #854d0e; }
.stat-pass    { background: #dcfce7; color: #166534; }
.stat-none    { background: #f1f5f9; color: #64748b; }

/* Main layout */
.main { display: flex; flex: 1; overflow: hidden; }

/* Left sidebar */
.sidebar {
  width: 400px; min-width: 320px; flex-shrink: 0;
  display: flex; flex-direction: column;
  border-right: 1px solid #e2e8f0; background: #fff;
}
.sidebar .search {
  padding: 14px 16px; border-bottom: 1px solid #f1f5f9;
}
.sidebar .search input {
  width: 100%; padding: 9px 14px;
  border: 1px solid #e2e8f0; border-radius: 10px;
  font-size: 13px; outline: none; background: #f8fafc;
  transition: border-color 0.15s, box-shadow 0.15s;
}
.sidebar .search input:focus {
  border-color: #2563eb; background: #fff;
  box-shadow: 0 0 0 3px rgba(37,99,235,0.12);
}
.sidebar .filter-chips {
  display: flex; gap: 6px; padding: 0 16px 12px; flex-wrap: wrap;
}
.sidebar .filter-chips button {
  font-size: 11px; padding: 4px 10px; border-radius: 9999px;
  border: 1px solid #e2e8f0; background: #f8fafc; color: #475569;
  cursor: pointer; font-weight: 600;
  transition: all 0.15s;
}
.sidebar .filter-chips button.active {
  background: #1e293b; color: #fff; border-color: #1e293b;
}
.sidebar .filter-chips button:hover:not(.active) { background: #eff6ff; }

.table-list { flex: 1; overflow-y: auto; padding: 4px 8px 12px; }
.tl-item {
  display: flex; align-items: center; gap: 10px;
  padding: 9px 12px; margin: 2px 0;
  cursor: pointer; border-radius: 10px;
  transition: background 0.1s;
  border: 1px solid transparent;
}
.tl-item:hover { background: #eff6ff; }
.tl-item.hovering { background: #fef3c7; border-color: #fde68a; }
.tl-item.active { background: #dbeafe; border-color: #93c5fd; }
.tl-item .tl-idx {
  font-size: 11px; color: #94a3b8; font-weight: 700;
  min-width: 24px; text-align: right;
}
.tl-item .tl-title {
  flex: 1; font-size: 13px; line-height: 1.35; color: #334155;
  overflow: hidden; text-overflow: ellipsis;
  display: -webkit-box; -webkit-line-clamp: 2; -webkit-box-orient: vertical;
}
.tl-item .tl-badge {
  font-size: 11px; padding: 2px 8px; border-radius: 9999px;
  white-space: nowrap; font-weight: 600; flex-shrink: 0;
}
.badge-major  { background: #fee2e2; color: #991b1b; }
.badge-manual { background: #fef9c3; color: #854d0e; }
.badge-pass   { background: #dcfce7; color: #166534; }
.badge-none   { background: #f1f5f9; color: #94a3b8; }

.empty { padding: 24px; text-align: center; color: #94a3b8; font-size: 13px; }

/* Right panel */
.content { flex: 1; display: flex; flex-direction: column; overflow: hidden; }
.content-header {
  display: flex; align-items: center; gap: 12px;
  padding: 12px 24px; background: #fff;
  border-bottom: 1px solid #e2e8f0; flex-shrink: 0;
}
.content-header .title { font-size: 14px; font-weight: 700; color: #1e293b; flex: 1; }
.content-header .pinned { font-size: 11px; padding: 2px 8px; border-radius: 9999px; background: #dbeafe; color: #1e40af; font-weight: 600; }
.toggle { display: inline-flex; background: #f1f5f9; border-radius: 10px; padding: 3px; gap: 2px; }
.toggle button {
  padding: 6px 14px; border: none; border-radius: 8px;
  background: transparent; color: #64748b; cursor: pointer;
  font-size: 12px; font-weight: 600;
  transition: all 0.15s;
}
.toggle button.active { background: #fff; color: #1e293b; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
.toggle button:disabled { opacity: 0.4; cursor: not-allowed; }

.content-body { flex: 1; overflow-y: auto; padding: 24px; }

.placeholder {
  display: flex; flex-direction: column; align-items: center; justify-content: center;
  height: 100%; color: #94a3b8; text-align: center;
}
.placeholder .icon { font-size: 56px; opacity: 0.4; margin-bottom: 14px; }
.placeholder p { font-size: 14px; line-height: 1.7; max-width: 380px; }
.placeholder .tip { font-size: 12px; margin-top: 8px; color: #cbd5e1; }

.report-shell { max-width: 1080px; }
.report-shell h1, .report-shell h2, .report-shell h3 { margin: 16px 0 10px; color: #0f172a; }
.report-shell h1 { font-size: 19px; }
.report-shell h2 { font-size: 16px; padding-bottom: 6px; border-bottom: 1px solid #e2e8f0; }
.report-shell h3 { font-size: 14px; }
.report-shell p { margin: 8px 0; line-height: 1.7; font-size: 13px; color: #334155; }
.report-shell ul, .report-shell ol { margin: 8px 0; padding-left: 24px; }
.report-shell li { line-height: 1.7; font-size: 13px; color: #334155; }
.report-shell blockquote {
  border-left: 3px solid #2563eb; padding: 8px 14px;
  margin: 10px 0; background: #eff6ff;
  font-size: 13px; border-radius: 0 8px 8px 0; color: #1e3a8a;
}
.report-shell code { background: #f1f5f9; padding: 2px 6px; border-radius: 4px; font-size: 12px; color: #0f172a; }
.report-shell pre { background: #0f172a; color: #e2e8f0; padding: 14px; border-radius: 10px; overflow-x: auto; font-size: 12px; margin: 12px 0; }
.report-shell hr { border: none; border-top: 1px solid #e2e8f0; margin: 18px 0; }
.report-shell a { color: #2563eb; text-decoration: none; }
.report-shell a:hover { text-decoration: underline; }

.report-shell table {
  border-collapse: collapse; width: 100%; margin: 12px 0; font-size: 12px;
  background: #fff; box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}
.report-shell th, .report-shell td {
  border: 1px solid #e2e8f0; padding: 6px 10px;
  text-align: left; vertical-align: top;
}
.report-shell th { background: #f8fafc; font-weight: 600; color: #475569; }
.report-shell tr:nth-child(even) td { background: #fafafa; }

/* Strip META lines that aren't useful in rendered view */
.report-shell h2:first-child + p:has(+ h2),
.report-shell p:has(>code:only-child) { font-family: ui-monospace, monospace; font-size: 11px; color: #94a3b8; }

/* Original xlsx table */
.xlsx-table {
  border-collapse: collapse; width: 100%; font-size: 12px;
  background: #fff; box-shadow: 0 1px 2px rgba(0,0,0,0.04);
}
.xlsx-table th, .xlsx-table td {
  border: 1px solid #e2e8f0; padding: 6px 10px;
  text-align: left; vertical-align: top;
}
.xlsx-table th { background: #dbeafe; color: #1e3a8a; font-weight: 600; }
.xlsx-table tr:nth-child(even) td { background: #f8fafc; }
.xlsx-meta {
  padding: 12px 16px; background: #fff;
  border: 1px solid #e2e8f0; border-radius: 10px;
  margin-bottom: 14px; font-size: 12px; color: #64748b;
}
.xlsx-meta strong { color: #1e293b; }

/* Scrollbar */
::-webkit-scrollbar { width: 8px; height: 8px; }
::-webkit-scrollbar-track { background: transparent; }
::-webkit-scrollbar-thumb { background: #cbd5e1; border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: #94a3b8; }

/* ═══════════════════════════════════════════════════════════════════════════
   Cover Page
   ═══════════════════════════════════════════════════════════════════════════ */
.cover-overlay {
  position: fixed; inset: 0; z-index: 9999;
  display: flex; align-items: center; justify-content: center;
  background: #fff;
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Microsoft YaHei", sans-serif;
  transition: opacity 0.5s, visibility 0.5s;
}
.cover-overlay.dismissed { opacity: 0; visibility: hidden; pointer-events: none; }

.cover-overlay::before {
  content: ''; position: absolute; top: 0; left: 0; right: 0; height: 320px;
  background: linear-gradient(180deg, #eff6ff 0%, rgba(239,246,255,0.3) 70%, transparent 100%);
  pointer-events: none;
}

.cover-card {
  position: relative; z-index: 1;
  text-align: center; max-width: 720px;
  padding: 48px 64px 60px;
}
.cover-logo {
  margin-bottom: 56px;
  display: inline-block;
}
.cover-logo img {
  height: 140px; width: auto;
  filter: drop-shadow(0 6px 16px rgba(37,99,235,0.08));
}
.cover-label {
  font-size: 13px; font-weight: 600; color: #2563eb;
  letter-spacing: 0.22em; text-transform: uppercase; margin-bottom: 24px;
}
.cover-h1 {
  font-size: 38px; font-weight: 800; color: #0f172a;
  line-height: 1.3; margin-bottom: 12px;
  letter-spacing: -0.015em;
}
.cover-h2 {
  font-size: 15px; color: #64748b; font-weight: 400; margin-bottom: 56px;
}

.cover-enter-btn {
  display: inline-flex; align-items: center; gap: 10px;
  padding: 16px 56px;
  background: #2563eb; color: #fff;
  border: none; border-radius: 14px;
  font-size: 16px; font-weight: 700; cursor: pointer;
  letter-spacing: 0.02em;
  transition: all 0.2s;
  box-shadow: 0 4px 20px rgba(37,99,235,0.30);
  font-family: inherit;
}
.cover-enter-btn:hover {
  background: #1d4ed8; transform: translateY(-1px);
  box-shadow: 0 8px 32px rgba(37,99,235,0.40);
}
.cover-enter-btn:active { transform: translateY(0); }

.cover-footer {
  margin-top: 48px; font-size: 12px; color: #94a3b8;
}
.cover-footer span { color: #64748b; font-weight: 500; }

/* ── Usage Tip ── */
.usage-tip {
  background: #eff6ff; padding: 12px 16px; border-radius: 12px;
  margin-bottom: 16px; font-size: 13px; color: #1e40af;
}
.usage-tip kbd {
  background: #dbeafe; padding: 1px 6px; border-radius: 3px;
}
"""


def build_html(items: list[dict], stats: dict, project_name: str) -> str:
    """Generate the full self-contained HTML page."""
    items_json = json.dumps(items, ensure_ascii=False)

    # ── Logo base64 ──
    logo_b64 = ""
    SCRIPT_DIR = Path(__file__).resolve().parent
    if _HAS_PIL:
        try:
            logo_path = SCRIPT_DIR.parent.parent.parent.parent / "frontend" / "public" / "logo.png"
            if logo_path.exists():
                logo = Image.open(str(logo_path))
                h = 120
                w = int(logo.width * h / logo.height)
                logo_small = logo.resize((w, h), Image.LANCZOS)
                buf = io.BytesIO()
                logo_small.save(buf, format="PNG", optimize=True)
                logo_b64 = "data:image/png;base64," + b64_mod.b64encode(buf.getvalue()).decode()
        except Exception:
            pass

    # Stats capsules — keep severity buckets in fixed order
    stat_parts = [f'<span class="stat-total">总计: {stats["total"]}</span>']
    for sev in ("major", "manual", "pass", "none"):
        n = stats.get(sev, 0)
        if n:
            label = SEVERITY_LABEL[sev].split(" ", 1)[-1]
            stat_parts.append(f'<span class="stat-{sev}">{label}: {n}</span>')
    stats_html = "\n    ".join(stat_parts)

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{html_mod.escape(project_name)} — 表格审查报告</title>
<style>{CSS}</style>
</head>
<body>

<!-- ═══════════════════════════════════════════════════════════════════════════
     Cover Page
     ═══════════════════════════════════════════════════════════════════════════ -->
<div class="cover-overlay" id="coverOverlay">
  <div class="cover-card">
    <div class="cover-logo">
      {f'<img src="{logo_b64}" alt="Logo" />' if logo_b64 else '''
      <svg width="88" height="88" viewBox="0 0 88 88" fill="none">
        <rect x="14" y="8" width="60" height="72" rx="10" fill="#eff6ff" stroke="#2563eb" stroke-width="2.5"/>
        <path d="M40 26v20m0 0l-8-8m8 8l8-8" stroke="#2563eb" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/>
        <circle cx="44" cy="58" r="12" fill="#2563eb" opacity="0.12" stroke="#2563eb" stroke-width="2"/>
        <path d="M39 58l3 3 7-7" stroke="#16a34a" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"/>
      </svg>'''}
    </div>

    <div class="cover-label">Clinical Trial Quality Control Report</div>

    <div class="cover-h1">表格内部一致性质控</div>
    <div class="cover-h2">Inner-Table Consistency Quality Control Report</div>

    <button class="cover-enter-btn" onclick="dismissCover()">
      <svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round">
        <polyline points="9 18 15 12 9 6"/>
      </svg>
      查看报告
    </button>

    <div class="cover-footer">
      由 <span>Inner QC Platform</span> 自动生成
    </div>
  </div>
</div>

<script>
function dismissCover() {{
  document.getElementById('coverOverlay').classList.add('dismissed');
}}
</script>

<header class="top-bar">
  {f'<img src="{logo_b64}" alt="Logo" style="height:28px;width:auto;" />' if logo_b64 else '''
  <svg viewBox="0 0 80 80" fill="none" width="28" height="28">
    <circle cx="40" cy="40" r="36" fill="#2563eb" opacity="0.12"/>
    <path d="M28 22h16l10 10v20a4 4 0 0 1-4 4H28a4 4 0 0 1-4-4V26a4 4 0 0 1 4-4z" fill="none" stroke="#2563eb" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/>
    <path d="M44 22v10h10" fill="none" stroke="#2563eb" stroke-width="3" stroke-linecap="round" stroke-linejoin="round"/>
    <circle cx="48" cy="52" r="10" fill="#2563eb" opacity="0.15" stroke="#2563eb" stroke-width="2"/>
    <path d="M44 52l3 2.5 5-5" fill="none" stroke="#2563eb" stroke-width="2.5" stroke-linecap="round" stroke-linejoin="round"/>
  </svg>'''}
  <h1>表格内部一致性质控报告</h1>
  <span class="subtitle">{html_mod.escape(project_name)}</span>
  <div class="stats">
    {stats_html}
  </div>
</header>

<div class="main">

  <aside class="sidebar">
    <div class="search">
      <input id="searchInput" type="text" placeholder="按表号或表名搜索…">
    </div>
    <div class="filter-chips" id="filterChips">
      <button data-filter="all" class="active">全部</button>
      <button data-filter="major">🔴 Major</button>
      <button data-filter="manual">🟡 待人工</button>
      <button data-filter="pass">🟢 核查无误</button>
      <button data-filter="none">⚪ 未核查</button>
    </div>
    <div id="tableList" class="table-list"></div>
  </aside>

  <section class="content">
    <div class="content-header">
      <div class="title" id="contentTitle">选择左侧表格查看审查结果</div>
      <span class="pinned" id="pinnedFlag" style="display:none;">📌 已固定</span>
      <div class="toggle" id="viewToggle">
        <button data-view="report" class="active">审查意见</button>
        <button data-view="table">原始表格</button>
      </div>
    </div>
    <div class="content-body" id="contentBody">
      <div class="placeholder">
        <div class="icon">📊</div>
        <p>悬停左侧表格即可预览审查报告</p>
        <p class="tip">点击表格固定显示 · 顶部切换"审查意见 / 原始表格" · ←↑↓→ 键盘导航 · Esc 取消固定</p>
      </div>
    </div>
  </section>

</div>

<script>
const ITEMS = {items_json};
const SEV_LABEL = {{
  major: '🔴 Major',
  manual: '🟡 待人工',
  pass: '🟢 核查无误',
  none: '⚪ 未核查',
}};

const tableList   = document.getElementById('tableList');
const contentBody = document.getElementById('contentBody');
const contentTitle= document.getElementById('contentTitle');
const pinnedFlag  = document.getElementById('pinnedFlag');
const searchInput = document.getElementById('searchInput');

let activeIdx = null;       // pinned index
let hoverIdx  = null;
let currentView = 'report'; // 'report' | 'table'
let currentFilter = 'all';

function render(filter, q) {{
  q = (q || '').toLowerCase().trim();
  tableList.innerHTML = '';
  const rows = ITEMS.filter(it => {{
    if (filter && filter !== 'all' && it.severity !== filter) return false;
    if (!q) return true;
    return (it.table_number || '').toLowerCase().includes(q)
        || (it.short_name || '').toLowerCase().includes(q)
        || String(it.index).includes(q);
  }});
  if (rows.length === 0) {{
    tableList.innerHTML = '<div class="empty">没有匹配的表格</div>';
    return;
  }}
  rows.forEach(it => {{
    const div = document.createElement('div');
    div.className = 'tl-item';
    if (it.index === activeIdx) div.classList.add('active');
    div.dataset.idx = it.index;
    div.innerHTML =
      '<span class="tl-idx">#' + String(it.index).padStart(2, '0') + '</span>' +
      '<span class="tl-title">' + escapeHtml(it.table_number) + ' ' + escapeHtml(it.short_name) + '</span>' +
      '<span class="tl-badge badge-' + it.severity + '">' + SEV_LABEL[it.severity] + '</span>';
    div.addEventListener('mouseenter', () => onHover(it.index));
    div.addEventListener('mouseleave', onLeave);
    div.addEventListener('click', () => onClick(it.index));
    tableList.appendChild(div);
  }});
}}

function onHover(idx) {{
  hoverIdx = idx;
  document.querySelectorAll('.tl-item').forEach(el => {{
    el.classList.toggle('hovering', parseInt(el.dataset.idx) === idx && idx !== activeIdx);
  }});
  showItem(idx, /*ephemeral*/ activeIdx === null || activeIdx !== idx);
}}

function onLeave() {{
  hoverIdx = null;
  document.querySelectorAll('.tl-item').forEach(el => el.classList.remove('hovering'));
  if (activeIdx !== null) showItem(activeIdx, false);
  else showPlaceholder();
}}

function onClick(idx) {{
  if (activeIdx === idx) {{
    activeIdx = null;
    showPlaceholder();
  }} else {{
    activeIdx = idx;
    showItem(idx, false);
  }}
  render(currentFilter, searchInput.value);
}}

function showItem(idx, ephemeral) {{
  const it = ITEMS.find(x => x.index === idx);
  if (!it) return;
  contentTitle.textContent = it.table_number + ' ' + it.short_name;
  pinnedFlag.style.display = (activeIdx === idx && !ephemeral) ? '' : 'none';

  // Disable "原始表格" toggle if we have no xlsx for this row
  const tableBtn = document.querySelector('#viewToggle button[data-view="table"]');
  tableBtn.disabled = !it.table_html;
  const reportBtn = document.querySelector('#viewToggle button[data-view="report"]');
  reportBtn.disabled = !it.report_html;

  let view = currentView;
  if (view === 'table' && !it.table_html) view = 'report';
  if (view === 'report' && !it.report_html) view = 'table';

  if (view === 'report' && it.report_html) {{
    contentBody.innerHTML = '<div class="report-shell">' + it.report_html + '</div>';
  }} else if (view === 'table' && it.table_html) {{
    const meta = '<div class="xlsx-meta">' +
      '<strong>' + escapeHtml(it.table_number) + '</strong> ' + escapeHtml(it.short_name) +
      ' · 文件: <code>' + escapeHtml(it.filename || '') + '</code>' +
      '</div>';
    contentBody.innerHTML = meta + it.table_html;
  }} else {{
    contentBody.innerHTML = '<div class="placeholder"><div class="icon">📭</div>' +
      '<p>该表格' + (view === 'report' ? '没有审查报告' : '没有原始表格文件') + '</p></div>';
  }}
}}

function showPlaceholder() {{
  contentTitle.textContent = '选择左侧表格查看审查结果';
  pinnedFlag.style.display = 'none';
  contentBody.innerHTML =
    '<div class="placeholder"><div class="icon">📊</div>' +
    '<p>悬停左侧表格即可预览审查报告</p>' +
    '<p class="tip">点击表格固定显示 · 顶部切换"审查意见 / 原始表格" · Esc 取消固定</p></div>';
}}

document.getElementById('viewToggle').addEventListener('click', e => {{
  if (e.target.tagName !== 'BUTTON' || e.target.disabled) return;
  currentView = e.target.dataset.view;
  document.querySelectorAll('#viewToggle button').forEach(b =>
    b.classList.toggle('active', b.dataset.view === currentView));
  const target = activeIdx !== null ? activeIdx : hoverIdx;
  if (target !== null) showItem(target, activeIdx !== target);
}});

document.getElementById('filterChips').addEventListener('click', e => {{
  if (e.target.tagName !== 'BUTTON') return;
  currentFilter = e.target.dataset.filter;
  document.querySelectorAll('#filterChips button').forEach(b =>
    b.classList.toggle('active', b.dataset.filter === currentFilter));
  render(currentFilter, searchInput.value);
}});

searchInput.addEventListener('input', () => render(currentFilter, searchInput.value));

document.addEventListener('keydown', e => {{
  if (e.key === 'Escape' && activeIdx !== null) {{
    activeIdx = null;
    showPlaceholder();
    render(currentFilter, searchInput.value);
  }}
  if (e.key === 'ArrowDown' || e.key === 'ArrowUp') {{
    e.preventDefault();
    const items = Array.from(document.querySelectorAll('.tl-item'));
    if (items.length === 0) return;
    const currentIdx = items.findIndex(el => el.classList.contains('active'));
    let nextIdx;
    if (e.key === 'ArrowDown') nextIdx = currentIdx < 0 ? 0 : Math.min(currentIdx + 1, items.length - 1);
    else nextIdx = currentIdx < 0 ? items.length - 1 : Math.max(currentIdx - 1, 0);
    const nextItem = items[nextIdx];
    const tableIdx = parseInt(nextItem.dataset.idx);
    onClick(tableIdx);
    nextItem.scrollIntoView({{ block: 'nearest' }});
  }}
}});

function escapeHtml(s) {{
  const d = document.createElement('div');
  d.textContent = String(s == null ? '' : s);
  return d.innerHTML;
}}

render('all', '');
</script>

</body>
</html>"""


# ── Main ────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(
        description="将 subagent 输出的 markdown 审查报告合成为单页 HTML",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--reports-dir", required=True,
                    help="包含 qc_NN.md / qc_NN.json 的目录")
    ap.add_argument("--tables-dir", required=True,
                    help="包含 NN-*.xlsx 原始表格的目录")
    ap.add_argument("--output", default=None,
                    help="输出 HTML 路径（默认为 reports-dir 同级的 report-viewer.html）")
    ap.add_argument("--project-name", default=None,
                    help="项目名（用于页面标题；默认取 reports-dir 上一级目录名）")
    args = ap.parse_args()

    reports_dir = Path(args.reports_dir).expanduser().resolve()
    tables_dir  = Path(args.tables_dir).expanduser().resolve()
    if not reports_dir.is_dir():
        ap.error(f"reports-dir 不存在: {reports_dir}")
    if not tables_dir.is_dir():
        ap.error(f"tables-dir 不存在: {tables_dir}")

    output_path = (Path(args.output).expanduser().resolve()
                   if args.output else reports_dir.parent / "report-viewer.html")
    project_name = args.project_name or reports_dir.parent.name

    print(f"📝 读取审查报告: {reports_dir}")
    reports = load_reports(reports_dir)
    print(f"   → {len(reports)} 份报告")

    print(f"📊 读取原始表格: {tables_dir}")
    tables = load_tables(tables_dir)
    print(f"   → {len(tables)} 张表格")

    # Merge by index — union of all indices present in either dir
    all_idx = sorted(set(reports.keys()) | set(tables.keys()))
    items: list[dict] = []
    stats = {"total": 0, "major": 0, "manual": 0, "pass": 0, "none": 0}

    for idx in all_idx:
        rep = reports.get(idx)
        tab = tables.get(idx)

        # Prefer table info for display title, fall back to report
        if tab:
            table_number = tab["table_number"]
            short_name   = tab["short_name"]
        elif rep and rep["title"]:
            tm = re.search(r"(表\s*\d+(?:\.\d+)*)\s*(.*)", rep["title"])
            if tm:
                table_number = tm.group(1).strip()
                short_name   = tm.group(2).strip()
            else:
                table_number = f"表 {idx}"
                short_name   = rep["title"]
        else:
            table_number = f"表 {idx}"
            short_name   = "(无标题)"

        severity = rep["severity"] if rep else "none"
        stats[severity] = stats.get(severity, 0) + 1
        stats["total"] += 1

        items.append({
            "index": idx,
            "table_number": table_number,
            "short_name": short_name,
            "filename": tab["filename"] if tab else None,
            "severity": severity,
            "conclusion": rep["conclusion"] if rep else "未核查",
            "pending": rep["pending"] if rep else 0,
            "report_html": rep["html"] if rep else None,
            "table_html": tab["html"] if tab else None,
        })

    print(f"📈 统计: major={stats['major']}, 待人工={stats['manual']}, "
          f"核查无误={stats['pass']}, 未核查={stats['none']}")

    print(f"🔨 生成 HTML…")
    html = build_html(items, stats, project_name)
    output_path.write_text(html, encoding="utf-8")
    size_kb = output_path.stat().st_size / 1024
    print(f"✅ 完成: {output_path} ({size_kb:.0f} KB)")


if __name__ == "__main__":
    main()
