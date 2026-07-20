# -*- coding: utf-8 -*-
"""inner-qc 共享确定性核查库（单一真源）。

设计原则：**算术核查交给代码，不交给 LLM 眼算。**
每个 subagent 只做语义工作（读 xlsx → 定位单元格），所有求和/百分比/次序/边界
判断一律 import 本模块的函数完成；问题清单用 Issues 收集并落盘为 JSON，供
Phase 6 build_reports.py 确定性合并。

这样替代了 7 个规则文档里各自复制的 to_num/check_* 片段，消除实现漂移。

用法（subagent 内）::

    import sys; sys.path.insert(0, "<skill>/scripts")
    from qc_lib import (read_grid, to_num, parse_n_pct, parse_n_missing,
                        lead_int, find_N, check_sum, check_pct, check_ordered,
                        check_le, check_equal, pct_tol, Issues)

    grid = read_grid(XLSX_PATH)          # 二维 str，等价旧模板的 read_csv 结果
    iss  = Issues(table_index="14", table_type="标准定性定量表",
                  title="表 14.4.1 疗效分级", analysis_set="FAS",
                  n_by_group={"试验组": 50, "对照组": 48, "合计": 98})
    ...  # 定位单元格后调用 check_* 并 iss.add(...)
    iss.to_json("qc_output/qc_14.json")
"""
import csv
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # 仅在没装 openpyxl 时触发；CSV 输入不需要它
    load_workbook = None

# ------------------------------------------------------------------
# 正则与基础解析
# ------------------------------------------------------------------
NUM = re.compile(r"-?\d+(?:\.\d+)?")
N_PCT = re.compile(r"(\d+)\s*[（(]\s*([\d.]+)\s*%")    # "24(48.0%)" -> (24, 48.0)
N_MISS = re.compile(r"(\d+)\s*[（(]\s*(\d+)\s*[)）]")  # "49(1)"     -> (49, 1)
NEQ = re.compile(r"N\s*[=＝]\s*(\d+)")


def to_num(x):
    """从任意单元格取第一个数字（容错全角 ％）。取不到返回 None。"""
    if x is None:
        return None
    m = NUM.search(str(x).replace("％", "%"))
    return float(m.group()) if m else None


def parse_n_pct(cell):
    """'24(48.0%)' -> (24, 48.0)；取不到返回 (None, None)。"""
    m = N_PCT.search(str(cell))
    return (int(m.group(1)), float(m.group(2))) if m else (None, None)


def parse_n_missing(cell):
    """'49(1)' -> (49, 1)；取不到返回 (None, None)。"""
    m = N_MISS.search(str(cell))
    return (int(m.group(1)), int(m.group(2))) if m else (None, None)


def lead_int(cell):
    """'47(94%)' -> 47（取前导整数）；取不到返回 None。"""
    m = re.match(r"\s*(\d+)", str(cell))
    return int(m.group(1)) if m else None


def find_N(text):
    """从表头/表题里抓 'N=98' 的 98；取不到返回 None。"""
    m = NEQ.search(str(text))
    return int(m.group(1)) if m else None


# ------------------------------------------------------------------
# 确定性核查（返回 True/False/None；None = 数据缺失，应静默跳过）
# ------------------------------------------------------------------
def check_sum(total, parts, tol=0):
    """total == Σparts ?（计数求和默认 tol=0）。"""
    t = to_num(total)
    ns = [to_num(p) for p in parts]
    if t is None or any(n is None for n in ns):
        return None
    return abs(t - sum(ns)) <= tol


def check_pct(printed_pct, n, N, tol=0.15):
    """印刷百分比 == n/N*100 ?（默认容差 0.15 个百分点，吸收四舍五入）。"""
    p = to_num(printed_pct)
    if p is None or not N:
        return None
    return abs(p - n / N * 100) <= tol


def check_ordered(seq):
    """seq 是否非降序（min<=Q1<=中位<=Q3<=max）。"""
    ns = [to_num(v) for v in seq]
    if any(n is None for n in ns):
        return None
    return all(ns[i] <= ns[i + 1] for i in range(len(ns) - 1))


def check_le(a, b):
    """a <= b ?（基准比对统一用 <=，子集/亚组人数更少属正常）。"""
    a, b = to_num(a), to_num(b)
    if a is None or b is None:
        return None
    return a <= b


def check_equal(values, tol=0):
    """所有 values 相等 ?（能转数字按数值比，否则按去空白字符串比）。"""
    ns = [to_num(v) for v in values]
    if all(n is not None for n in ns):
        return max(ns) - min(ns) <= tol
    return len({str(v).strip() for v in values}) <= 1


def pct_tol(n_levels):
    """Σ%≈100 的容差：类别越多容差越大。"""
    return 0.05 * n_levels + 0.35


# ------------------------------------------------------------------
# 表格读取（xlsx / csv 同一接口，替代旧模板的 pd.read_csv）
# ------------------------------------------------------------------
def read_grid(path):
    """读 xlsx 或 csv 为二维字符串列表（空单元格 -> ""）。

    解决旧规则模板写 `pd.read_csv(CSV_PATH)` 而提取脚本实际产出 .xlsx 的错位。
    """
    p = str(path)
    if p.lower().endswith((".xlsx", ".xlsm")):
        if load_workbook is None:
            raise RuntimeError("读取 xlsx 需要 openpyxl：pip install openpyxl")
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        grid = [["" if c is None else str(c) for c in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
        return grid
    with open(p, newline="", encoding="utf-8-sig") as f:
        return [[("" if c is None else str(c)) for c in row] for row in csv.reader(f)]


# ------------------------------------------------------------------
# 问题清单收集 + 结构化落盘
# ------------------------------------------------------------------
class Issues:
    """收集单张表的核查问题，输出结构化 JSON（供 build_reports.py 合并）。

    level 取值：CRITICAL / MAJOR / MINOR / SUGGESTION / 待人工。
    通过项默认不记录（保持清单干净）；结论 = 所有硬问题的最高级别，无硬问题则 PASS。
    """

    ORDER = ["SUGGESTION", "MINOR", "MAJOR", "CRITICAL"]  # 由低到高

    def __init__(self, table_index, table_type, title,
                 analysis_set=None, n_by_group=None):
        self.meta = {
            "table_index": str(table_index),
            "table_type": table_type,
            "title": title,
            "analysis_set": analysis_set,        # 表题声明的分析集 FAS/PPS/SS
            "n_by_group": n_by_group or {},      # {组名: N}，供跨表 R-021
        }
        self.items = []

    def add(self, rule, level, where, expected, found, note=""):
        self.items.append({
            "rule": rule, "level": level, "where": where,
            "expected": expected, "found": found, "note": note,
        })

    def conclusion(self):
        hard = [it["level"] for it in self.items if it["level"] in self.ORDER]
        if not hard:
            return "PASS"
        return max(hard, key=self.ORDER.index)

    def pending_count(self):
        return sum(1 for it in self.items if it["level"] == "待人工")

    def to_json(self, out_path):
        payload = {
            "meta": self.meta,
            "conclusion": self.conclusion(),
            "pending": self.pending_count(),
            "issues": self.items,
        }
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(payload, ensure_ascii=False, indent=2),
                       encoding="utf-8")
        return payload


if __name__ == "__main__":
    # 冒烟自检
    assert check_sum(98, [50, 48]) is True
    assert check_sum(98, [50, 47]) is False
    assert check_pct(48.0, 24, 50) is True
    assert check_pct(52.0, 25, 50) is False
    assert check_ordered([1, 2, 3, 3, 5]) is True
    assert check_ordered([1, 9, 3]) is False
    assert check_le(50, 98) is True and check_le(99, 98) is False
    assert parse_n_pct("24(48.0%)") == (24, 48.0)
    assert parse_n_missing("49(1)") == (49, 1)
    assert lead_int("47(94%)") == 47
    assert find_N("试验组(N=50)") == 50
    print("qc_lib 自检通过 ✓")
