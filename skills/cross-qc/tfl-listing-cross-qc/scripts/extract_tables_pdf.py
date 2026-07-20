#!/usr/bin/env python3
"""
从 PDF 中提取所有表格/清单为 Excel 文件，遵循 tfl-listing-cross-qc 命名规范。

用法:
    # 双文件模式（自动检测类型）
    python3 extract_tables_pdf.py <表格.pdf> <清单.pdf> [输出目录]

    # 单文件模式
    python3 extract_tables_pdf.py <文件.pdf> [--out 输出目录] [--type 表格|清单]

命名规则: {编号:02d}-{标题}.xlsx
跨页同名表格/清单自动合并（仅相邻且列标题行完全一致时）
"""

import pdfplumber
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, re, sys, json
from datetime import datetime
from collections import Counter


def clean(v):
    if v is None: return ""
    return re.sub(r'\s+', ' ', str(v).replace('\n', ' ').replace('\r', ' ')).strip()


def sanitize_filename(name):
    name = re.sub(r'[/\\:*?"<>|]', '-', name)
    name = re.sub(r'\s+', ' ', name).strip()
    if len(name) > 120: name = name[:120]
    return name


def is_toc_page(page, doc_type):
    """判断是否为目录页"""
    t = page.extract_text()
    if not t: return False
    lines = t.split('\n')
    if '目录' in lines[:3]:
        return True
    prefix = '清单' if doc_type == '清单' else '表'
    toc = sum(1 for l in lines if re.match(rf'^({prefix})\s*\d', l.strip()))
    sep = sum(1 for l in lines if '........' in l or '……' in l)
    return (toc > 3 and sep > 3) if doc_type == '清单' else (toc > 5 and sep > 3)


def extract_title(page, doc_type):
    """提取标题：表格用 '表 X.X.X.X ...' 模式，清单用 '清单 N ...' 模式"""
    text = page.extract_text()
    if not text: return None
    if doc_type == '清单':
        pattern = r'^(清单\s*\d+\s+.+)'
        exclude = r'^清单\s*\d+\s*$'
    else:
        pattern = r'^(表\s*\d[\d.]*\s+.+)'
        exclude = r'^表\s*\d[\d.]*\s*$'

    for line in text.split('\n'):
        line = line.strip()
        m = re.match(pattern, line)
        if not m: continue
        title = m.group(1).strip()
        # 排除仅编号无实质内容
        if re.match(exclude, title): continue
        # 排除统计方法/样板行（仅表格）
        if doc_type == '表格' and re.search(
            r'(统计方法|检验统计量|P值\s*$|例数\(缺失\)|成功\s*n\(%|失败\s*n\(%|男\s*n\(%|女\s*n\()', title):
            continue
        return title
    return None


def is_data_table(t):
    if not t or len(t) < 2: return False
    if len(t[0]) <= 1: return False
    return sum(1 for r in t[1:] if any(clean(c) for c in r)) >= 2


def col_count(t):
    return len(t[0]) if t else 0


def header_signature(table):
    """仅比较第一行列标题——用于判断跨页续表"""
    return "|".join(clean(c) for c in table[0]) if table else ""


def merge_fragments(tables_raw):
    """同页碎片合并：列数相同则追加，否则新表格"""
    merged = []
    cur = None
    for t in tables_raw:
        if not t or len(t) < 2: continue
        if not cur:
            cur = [list(r) for r in t]; continue
        if col_count(cur) == col_count(t) and col_count(cur) > 1:
            cur.extend([list(r) for r in t])
        elif col_count(cur) <= 1 and col_count(t) > 1:
            cur = [list(r) for r in t]
        else:
            merged.append(cur); cur = [list(r) for r in t]
    if cur: merged.append(cur)
    return merged


# ============================================================
# 质量检查
# ============================================================

def check_quality(table_data, doc_type):
    issues, score = [], 100
    if not table_data:
        return [("致命", "表格数据为空")], 0
    n_rows, n_cols = len(table_data), len(table_data[0]) if table_data else 0

    if n_cols <= 1:
        return [("严重", "列数过少")], 0

    # 多级表头检测
    has_mh = False
    if len(table_data) >= 3:
        h0 = sum(1 for c in table_data[0] if clean(c) == "")
        h1 = sum(1 for c in table_data[1] if clean(c) == "")
        has_mh = h0 > 0 and h1 > 0

    total = sum(len(r) for r in table_data)
    empty = sum(1 for r in table_data for c in r if clean(c) == "")
    ratio = empty / total if total else 1

    if ratio > 0.75:
        issues.append(("警告", f"空值率{ratio:.0%}"))
        score -= 25
    elif ratio > 0.5:
        issues.append(("信息", f"空值率{ratio:.0%}（多级表头或稀疏列）"))
        score -= 10

    if n_rows < 3:
        issues.append(("警告", "数据行不足3行"))
        score -= 15

    # 重复行检测（跳过统计样板行）
    if n_rows > 3 and doc_type == '表格':
        BOILERPLATE = {'统计方法', '检验统计量', 'P值', '例数(缺失)', '合计',
                       'Fisher精确检验', '卡方检验', 'Wilcoxon秩和检验', 't检验',
                       'CMH检验', '校正卡方检验', '-', '—', '0.000', '1.000'}
        sigs = []
        for r in table_data[1:]:
            cl = [clean(c) for c in r]
            ne = [c for c in cl if c]
            if len(ne) == 1 and ne[0] in BOILERPLATE: continue
            sigs.append("|".join(cl))
        dup = len(sigs) - len(set(sigs))
        if dup > 3:
            issues.append(("警告", f"{dup}行重复（统计样板行）"))
            score -= min(20, dup * 3)

    if n_cols > 25:
        issues.append(("警告", f"列数{n_cols}，检查是否含非表格数据"))
        score -= 10

    return issues, max(0, score)


# ============================================================
# Excel 导出
# ============================================================

HEADER_FONT = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
DATA_FONT = Font(name='微软雅黑', size=10)
DATA_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
NUM_ALIGN = Alignment(horizontal='center', vertical='center')
BORDER = Border(
    left=Side('thin', 'B0B0B0'), right=Side('thin', 'B0B0B0'),
    top=Side('thin', 'B0B0B0'), bottom=Side('thin', 'B0B0B0'))
ALT_FILL = PatternFill(start_color='EBF1F8', end_color='EBF1F8', fill_type='solid')
SECTION_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
SECTION_FONT = Font(name='微软雅黑', bold=True, size=10, color='1F3864')


def is_num(v):
    try:
        float(v.replace('%', '').replace(',', '').replace('±', '').strip('() '))
        return True
    except:
        return False


def save_excel(table_data, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    multi_header = False
    if len(table_data) >= 3:
        h0 = sum(1 for c in table_data[0] if clean(c) == "")
        h1 = sum(1 for c in table_data[1] if clean(c) == "")
        multi_header = h0 > 0 and h1 > 0

    for ri, row in enumerate(table_data):
        is_hdr = ri < (3 if multi_header else 1)
        is_sec = (len([c for c in row if clean(c)]) == 1 and len(row) > 1
                  and ri > 0 and not is_hdr)

        for ci, cell in enumerate(row):
            val = clean(cell)
            xc = ws.cell(row=ri + 1, column=ci + 1, value=val)
            xc.border = BORDER
            if is_hdr:
                xc.font = HEADER_FONT
                xc.fill = HEADER_FILL
                xc.alignment = HEADER_ALIGN
            elif is_sec:
                xc.font = SECTION_FONT
                xc.fill = SECTION_FILL
                xc.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            else:
                xc.font = DATA_FONT
                xc.alignment = NUM_ALIGN if is_num(val) else DATA_ALIGN
                if ri % 2 == 0:
                    xc.fill = ALT_FILL

    for ci in range(len(table_data[0]) if table_data else 0):
        mw = 0
        for row in table_data:
            if ci < len(row):
                v = str(clean(row[ci]))
                mw = max(mw, min(len(v) + len(re.findall(r'[一-鿿]', v)), 60))
        ws.column_dimensions[get_column_letter(ci + 1)].width = max(mw + 4, 8)

    if len(table_data) > 1:
        ws.freeze_panes = 'A2' if not multi_header else 'A4'
        ws.auto_filter.ref = f"A1:{get_column_letter(len(table_data[0]))}{len(table_data)}"

    wb.save(path)


# ============================================================
# 核心提取逻辑
# ============================================================

def process_pdf(pdf_path, output_dir, doc_type):
    """处理单个 PDF，提取所有表格/清单并导出 Excel"""
    if not os.path.exists(pdf_path):
        print(f"⚠ 文件不存在，跳过: {pdf_path}")
        return

    print(f"\n处理: {pdf_path}  (类型: {doc_type})")

    pdf = pdfplumber.open(pdf_path)
    total = len(pdf.pages)
    print(f"  PDF 页数: {total}")

    # ==== 第一遍：收集所有页面 ====
    page_info = []  # [(page_num, title, [tables])]

    for pn, page in enumerate(pdf.pages, 1):
        if is_toc_page(page, doc_type):
            page_info.append((pn, "__TOC__", []))
            continue

        tables_raw = page.extract_tables() or []
        merged = merge_fragments(tables_raw)
        real = [t for t in merged if is_data_table(t)]
        title = extract_title(page, doc_type)
        page_info.append((pn, title, real))

    # ==== 第二遍：跨页合并（仅相邻且列标题签名一致） ====
    merged_tables = []  # [(title, table, first_page, last_page, signature)]
    current = None

    for pn, title, tables in page_info:
        if title == "__TOC__":
            if current is not None:
                merged_tables.append(current)
                current = None
            continue
        if not tables:
            continue

        for ti, tdata in enumerate(tables):
            if title and ti == 0:
                t_title = title
            elif title and ti > 0:
                t_title = f"{title}（续{ti}）"
            elif current is not None:
                t_title = current[0]
            else:
                t_title = f"未命名{'清单' if doc_type == '清单' else '表格'}"

            t_sig = header_signature(tdata)

            if (current is not None and t_sig == current[4]
                    and col_count(tdata) == col_count(current[1])):
                # 列标题一致 + 列数相同 → 跨页合并
                current = (current[0],
                           current[1] + [list(r) for r in tdata],
                           current[2], pn,
                           current[4])
            else:
                if current is not None:
                    merged_tables.append(current)
                current = (t_title, [list(r) for r in tdata], pn, pn, t_sig)

    if current is not None:
        merged_tables.append(current)

    # ==== 去重 + 排序 ====
    titles = [t[0] for t in merged_tables]
    title_counts = Counter(titles)
    seen = {}
    final_tables = []

    for title, tdata, first_p, last_p, sig in merged_tables:
        if title_counts[title] > 1:
            n = seen.get(title, 0) + 1
            seen[title] = n
            if first_p == last_p:
                unique = f"{title}（第{n}部分）"
            else:
                unique = f"{title}（p{first_p}-{last_p}，第{n}部分）"
        else:
            if first_p != last_p:
                unique = f"{title}（p{first_p}-{last_p}）"
            else:
                unique = title
        final_tables.append((unique, tdata, first_p, last_p))

    # 按编号排序
    def sort_key(item):
        t, _, _, _ = item
        if doc_type == '清单':
            m = re.match(r'清单\s*(\d+)', t)
        else:
            m = re.match(r'表\s*([\d.]+)', t)
        if m:
            parts = m.group(1).split('.')
            return tuple(int(p) for p in parts)
        return (9999,)

    final_tables.sort(key=sort_key)

    print(f"  合并后: {len(final_tables)} 个{'清单' if doc_type == '清单' else '表格'}")

    # ==== 输出 ====
    os.makedirs(output_dir, exist_ok=True)
    results = []

    for seq, (title, tdata, first_p, last_p) in enumerate(final_tables, 1):
        issues, score = check_quality(tdata, doc_type)
        if score >= 90: grade = 'A'
        elif score >= 75: grade = 'B'
        elif score >= 60: grade = 'C'
        elif score >= 40: grade = 'D'
        else: grade = 'F'

        safe = sanitize_filename(title)
        fname = f"{seq:02d}-{safe}.xlsx"
        fpath = os.path.join(output_dir, fname)

        try:
            save_excel(tdata, fpath)
            n_rows = len(tdata)
            n_cols = len(tdata[0]) if tdata else 0
            pinfo = f"p{first_p}" if first_p == last_p else f"p{first_p}-{last_p}"
            print(f"  → {fname}  ({n_rows}行×{n_cols}列  {pinfo})")
        except Exception as e:
            print(f"  ✗ 保存失败: {e}")
            issues.append(("严重", str(e)))

        for sev, msg in issues:
            print(f"    [{sev}] {msg}")

        results.append({
            'seq': seq, 'title': title, 'file': fname,
            'pages': f"{first_p}-{last_p}" if first_p != last_p else str(first_p),
            'rows': n_rows, 'cols': n_cols,
            'score': score, 'grade': grade, 'issues': issues
        })

    pdf.close()

    # 质量报告
    grades = Counter(r['grade'] for r in results)
    avg = sum(r['score'] for r in results) / len(results) if results else 0
    problems = [r for r in results if r['score'] < 75]

    report = []
    report.append(f"# {'清单' if doc_type == '清单' else '表格'}提取质量检查报告\n")
    report.append(f"**时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report.append(f"**源文件**: `{os.path.basename(pdf_path)}`")
    report.append(f"**命名规范**: `{{编号:02d}}-{{标题}}.xlsx`\n")
    report.append("---\n## 📊 提取概览\n")
    report.append(f"| 指标 | 数值 |")
    report.append(f"|------|------|")
    report.append(f"| PDF 总页数 | {total} |")
    report.append(f"| 提取总数 | {len(results)} |")
    report.append(f"| 平均质量 | {avg:.1f}/100 |")
    for g, desc in [('A', '优秀 ≥90'), ('B', '良好 ≥75'), ('C', '一般 ≥60'),
                    ('D', '较差 ≥40'), ('F', '很差 <40')]:
        report.append(f"| {g}级 ({desc}) | {grades.get(g, 0)} |")
    report.append(f"| ⚠ 需复核 (<75分) | {len(problems)} |\n")

    report.append("---\n## 📑 文件清单\n\n")
    report.append(f"| 序号 | 页码 | 等级 | 分数 | 行×列 | 文件名 |")
    report.append(f"|------|------|------|------|-------|--------|")
    for r in results:
        gi = {'A': '🟢', 'B': '🟡', 'C': '🟠', 'D': '🔴', 'F': '⛔'}.get(r['grade'], '⚪')
        report.append(
            f"| {r['seq']:02d} | {r['pages']} | {gi} {r['grade']} | {r['score']} | {r['rows']}×{r['cols']} | `{r['file']}` |")

    if problems:
        report.append("\n---\n## ⚠️ 需人工复核\n")
        for t in problems:
            mi = t['issues'][0][1] if t['issues'] else ''
            report.append(f"- `{t['file']}` (p{t['pages']}, {t['score']}分): {mi}")

    report_path = os.path.join(output_dir, "质量检查报告.md")
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report))

    meta = {
        'source': pdf_path, 'total_pages': total, 'total_items': len(results),
        'quality_avg': round(avg, 1), 'grades': dict(grades),
        'problems': len(problems), 'time': datetime.now().isoformat(),
        'tables': [{k: v for k, v in r.items() if k != 'issues'} for r in results]
    }
    with open(os.path.join(output_dir, 'metadata.json'), 'w', encoding='utf-8') as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

    print(f"  质量报告: {report_path}\n")
    print(f"  已保存到: {output_dir}\n")


# ============================================================
# CLI
# ============================================================

def detect_doc_type(filename):
    """从文件名推断文档类型（与 docx 版 extract_tables.py 一致）"""
    basename = os.path.basename(filename)
    if "清单" in basename:
        return "清单"
    if "表格" in basename or "table" in basename.lower():
        return "表格"
    return "表格"  # 默认按表格处理


def main():
    args = sys.argv[1:]
    if not args:
        print("用法:")
        print("  python3 extract_tables_pdf.py <表格.pdf> <清单.pdf> [输出目录]")
        print("  python3 extract_tables_pdf.py <文件.pdf> [--out 输出目录] [--type 表格|清单]")
        sys.exit(1)

    output_base = os.getcwd()
    doc_type_override = None
    files = []

    i = 0
    while i < len(args):
        a = args[i]
        if a == "--out" and i + 1 < len(args):
            i += 1
            output_base = args[i]
        elif a == "--type" and i + 1 < len(args):
            i += 1
            doc_type_override = args[i]
        else:
            files.append(a)
        i += 1

    if not files:
        print("错误: 未指定输入文件")
        sys.exit(1)

    # 验证文件扩展名
    for f in files:
        if not f.lower().endswith('.pdf'):
            print(f"⚠ 非 PDF 文件，跳过: {f}")
            files.remove(f)

    if not files:
        print("错误: 无有效 PDF 文件")
        sys.exit(1)

    print("=" * 70)
    print("PDF 表格/清单提取 — tfl-listing-cross-qc Phase 2")
    print(f"  命名规范: {{编号:02d}}-{{标题}}.xlsx")
    print("=" * 70)

    # 双文件模式：自动分配 表格/ 和 清单/ 子目录
    if len(files) == 2 and doc_type_override is None:
        for pdf_path in files:
            basename = os.path.basename(pdf_path)
            doc_type = detect_doc_type(basename)
            out_dir = os.path.join(output_base, doc_type)
            process_pdf(pdf_path, out_dir, doc_type)
    else:
        # 单文件模式
        for pdf_path in files:
            doc_type = doc_type_override or detect_doc_type(os.path.basename(pdf_path))
            out_dir = os.path.join(output_base, doc_type) if len(files) > 1 else output_base
            process_pdf(pdf_path, out_dir, doc_type)

    print("完成。")


if __name__ == "__main__":
    main()
