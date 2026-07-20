# -*- coding: utf-8 -*-
"""对 tables_output 中的 xlsx 进行分类，并按 编号-类型-标题-分析集.xlsx 重命名。

分析集字段只填纯 acronym（FAS/ITT/mITT/PPS/SS）或"随机化人群"或"-"，
禁止写成"FAS人群""ITT人群"等带后缀形式——即便表题里出现"（FAS 人群）"，
extract_analysis_set 也只抽出"FAS"。

直接内嵌表型分类核心逻辑，不依赖外部的 similarity/llm_fallback。
"""

import json
import os
import re
import sys

from openpyxl import load_workbook

# ============================================================
# 分类核心逻辑（提取自 表型分类.py，去除外部依赖）
# ============================================================

NEQ = re.compile(r"N\s*[=＝]\s*(\d+)")

TIME_KW = ("治疗前", "治疗后", "基线", "访视", "筛选期", "筛选",
           "入院", "出院", "术前", "术后", "随访",
           "周", "天", "月", "年")


# 协方差相关关键词
COVARIANCE_KW = ("协方差", "修正均数", "最小二乘", "置信区间")


def _clean(s):
    return re.split(r"[\n/(（]", s or "")[0].strip()


def _flatten(rows):
    return " ".join(c for r in rows for c in r)


def classify(name, grid):
    """判型并返回类型名称。

    参数
    ----
    name : str  表格标题
    grid : list[list[str]]  从 xlsx 读取的二维数据

    返回
    ----
    type_name : str  表型 或 'other'
    """
    h0 = grid[0] if grid else []
    h1 = grid[1] if len(grid) > 1 else []
    flat = _flatten([h0, h1]) + " " + name
    h0_join = " ".join(h0)
    h1_join = " ".join(h1)

    # ① 协方差
    if any(k in name for k in COVARIANCE_KW):
        return "协方差"

    # ② 入组病例表（表题含「入组病例」关键词）
    if "入组病例" in name:
        return "入组病例表"

    # ③ 病例分布表
    if "入组" in h0_join and any(k in flat for k in ("脱落", "完成")):
        return "病例分布表"

    # ④ 人群划分表
    set_hits_h0 = sum(h0.count(s) for s in ("FAS", "PPS", "SS"))
    h1_has_N = bool(NEQ.search(h1_join))
    h1_has_indicator = any(k in h1_join for k in ("项目", "指标"))
    if len(grid) >= 2 and set_hits_h0 >= 3 and not h1_has_N and not h1_has_indicator:
        return "人群划分表"

    # ⑤ 交叉表（结构判定：h1 有"组别" + 出现时间关键词）
    if "组别" in h1_join and any(k in flat for k in TIME_KW):
        return "交叉表"

    # ⑥ 事件表
    head_join = h0_join + " " + h1_join
    if ("人数(%)" in h1_join
            or ("例次" in h1_join and "人数" in h1_join)
            or ("例次" in head_join and "例数" in head_join)):
        return "事件表"

    # ⑦ 标准定性定量表（纯结构判定，不依赖表名——表名往往与模板不一致）
    #    版式特征：左侧有「指标」列，自上而下逐行枚举统计描述符——
    #      · 例数(缺失) / 例数（缺失数）   每个指标的第一行，最稳的锚点
    #      · 连续型：均值(标准差)、中位数、第25/75%分位数、最小值,最大值
    #      · 分类型：每个类别一行「X n(%)」，如 24(48.0%)，末尾跟 统计方法/检验统计量/P值
    #    与交叉表（表头列是「组别」）、事件表（表头列是「例次/例数（%）」）的
    #    根本区别就在「指标」列＋上述统计量；故先用表头排除这两类，再认统计内容。
    head = _flatten(grid[:3])      # 表头区（含可能的两级表头）
    body = _flatten(grid)          # 全表文本；统计描述符集中在指标列，扫全表更稳
    is_cross_or_event = ("组别" in head) or ("例次" in head)
    has_indicator_col = "指标" in head
    # 例数(缺失)/例数（缺失数）——容忍全角/半角括号与多余空格
    has_n_missing = bool(re.search(r"例数\s*[（(]\s*缺失", body))
    cont_kw = ("均值", "标准差", "中位数", "分位数", "最小值", "最大值", "几何均值")
    cont_hits = sum(1 for k in cont_kw if k in body)
    has_npct = bool(re.search(r"\d\s*[（(]\s*\d+(?:\.\d+)?\s*%", body))   # 24(48.0%)
    has_test = any(k in body for k in ("统计方法", "检验统计量", "P值", "p值"))
    if not is_cross_or_event and (
        has_n_missing
        or cont_hits >= 2
        or (has_indicator_col and has_npct and has_test)
    ):
        return "标准定性定量表"

    # 额外启发：表题含"交叉"或"移位"关键词 → 交叉表
    # （移位表 = shift table，前后状态转移矩阵，与交叉表同型）
    if "交叉" in name or "移位" in name:
        return "交叉表"

    return "other"


# ============================================================
# 分析集推断
# ============================================================

# 以下表型固定标"随机化人群"（入组人群），不依赖表题/父节括号
RAND_TYPES = ("人群划分表", "病例分布表", "入组病例表")

# 表题/父节括号里能识别的分析集标签（长的写在前面，避免 ITT 抢先匹配 mITT）
SET_LABELS = ("mITT", "ITT", "FAS", "PPS", "SS")

# 匹配表题末尾括号：…（FAS） / …(PPS) / …（SS ） / …（mITT）/ …（ITT 人群）
# 容忍 acronym 后跟"人群"或"集"等后缀，只抽出纯 acronym
TAIL_SET_RE = re.compile(r"[（(]\s*(mITT|ITT|FAS|PPS|SS)\s*(?:人群|集)?\s*[)）]\s*$")
# 匹配父节标题里出现的（可能不在末尾）分析集标签
ANY_SET_RE = re.compile(r"[（(]\s*(mITT|ITT|FAS|PPS|SS)\s*(?:人群|集)?\s*[)）]")


def extract_analysis_set(title, parents, table_type):
    """根据 表型 + 表题 + 父节路径，确定该表所属分析集。

    返回值只会是纯 acronym（FAS/ITT/mITT/PPS/SS）、"随机化人群"或 "-"，
    不会带"人群""集"等后缀。

    规则（按顺序判定，命中即返回）：
    1. 人群划分表 / 病例分布表 / 入组病例表 → 随机化人群
    2. 表题末尾括号里出现 FAS/ITT/mITT/PPS/SS → 取之
    3. 父节标题（从最近的开始往上）括号里出现 FAS/ITT/mITT/PPS/SS → 取之
    4. 仍取不到 → "-"
    """
    if table_type in RAND_TYPES:
        return "随机化人群"

    if title:
        m = TAIL_SET_RE.search(title)
        if m:
            return m.group(1)

    # 父节从最近（最深层）向上找
    for pt in reversed(parents or []):
        m = ANY_SET_RE.search(pt)
        if m:
            return m.group(1)

    return "-"


# ============================================================
# 文件批量分类与重命名
# ============================================================

def extract_name_from_filename(filename):
    """从文件名提取表格标题，去掉编号前缀和 .xlsx 后缀。

    编号使用 \\d+ 匹配，兼容 01-99 的两位以及 100+ 的三位以上编号
    （extract_tables.py 用 {i+1:02d} 格式化，超过 99 后自然扩到 3 位）。
    """
    stem = filename.replace(".xlsx", "")
    # 匹配 "01-表 x.x.x.x 标题" / "100-表 ..." / "01-标题"
    m = re.match(r"^\d+-(.+)", stem)
    if m:
        return m.group(1)
    return stem


def load_tables_meta(input_dir):
    """读 extract_tables.py 写出的 tables_meta.json；不存在则返回 {}。"""
    meta_path = os.path.join(input_dir, "tables_meta.json")
    if not os.path.exists(meta_path):
        return {}
    with open(meta_path, encoding="utf-8") as f:
        return json.load(f)


def main():
    input_dir = sys.argv[1] if len(sys.argv) > 1 else "./tables_output"
    output_dir = sys.argv[2] if len(sys.argv) > 2 else input_dir  # 默认原地重命名

    if not os.path.isdir(input_dir):
        print(f"错误: 目录不存在: {input_dir}")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    # 收集 xlsx 文件
    files = sorted(
        [f for f in os.listdir(input_dir) if f.endswith(".xlsx") and not f.startswith("~")]
    )

    if not files:
        print(f"目录 {input_dir} 中未找到 xlsx 文件")
        sys.exit(1)

    meta = load_tables_meta(input_dir)

    print(f"找到 {len(files)} 个 xlsx 文件\n")
    print(f"{'编号':<6} {'类型':<14} {'分析集':<10} {'标题'}")
    print("-" * 90)

    results = []
    for fname in files:
        # 提取编号（支持 100+ 的三位以上编号；extract_tables.py 用 {i+1:02d}）
        m = re.match(r"^(\d+)-", fname)
        num = m.group(1) if m else "00"

        # 读取 xlsx 数据
        path = os.path.join(input_dir, fname)
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        grid = []
        for row in ws.iter_rows(values_only=True):
            grid.append([str(c) if c is not None else "" for c in row])
        wb.close()

        # 提取标题
        title = extract_name_from_filename(fname)

        # 分类
        tp = classify(title, grid)

        # 推断分析集（用 tables_meta.json 里的标题/父节，比文件名更可靠）
        meta_entry = meta.get(num, {})
        meta_title = meta_entry.get("title") or title
        parents = meta_entry.get("parents", [])
        analysis_set = extract_analysis_set(meta_title, parents, tp)

        print(f"{num:<6} {tp:<14} {analysis_set:<10} {title}")

        # 新文件名：编号-类型-标题-分析集.xlsx
        # 清理标题中的非法字符
        safe_title = re.sub(r'[/\\:*?"<>|]', "-", title)
        if len(safe_title) > 100:
            safe_title = safe_title[:100]
        new_name = f"{num}-{tp}-{safe_title}-{analysis_set}.xlsx"

        old_path = os.path.join(input_dir, fname)
        new_path = os.path.join(output_dir, new_name)

        if old_path == new_path:
            results.append((num, tp, analysis_set, title, "保持不变"))
        else:
            os.rename(old_path, new_path)
            results.append((num, tp, analysis_set, title, "已重命名"))

    print("-" * 90)

    # 统计
    type_counts = {}
    set_counts = {}
    for _, tp, ana, _, _ in results:
        type_counts[tp] = type_counts.get(tp, 0) + 1
        set_counts[ana] = set_counts.get(ana, 0) + 1

    print("\n分类统计:")
    for tp, cnt in sorted(type_counts.items(), key=lambda x: -x[1]):
        print(f"  {tp}: {cnt} 张")

    print("\n分析集分布:")
    for ana, cnt in sorted(set_counts.items(), key=lambda x: -x[1]):
        print(f"  {ana}: {cnt} 张")

    if output_dir != input_dir:
        print(f"\n输出目录: {output_dir}")
    print(f"\n共处理 {len(results)} 个文件。")


if __name__ == "__main__":
    main()
